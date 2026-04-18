"""
WellX Medical Insurance AI Underwriter
=======================================
A Streamlit web app that uses Anthropic Vision (Claude) to parse DHA reports
and generate insurance premium quotes following the DHA Report Analysis SOP.

Requirements:
    pip install streamlit anthropic pdf2image pillow pandas openpyxl

System dependency for pdf2image:
    macOS:   brew install poppler
    Ubuntu:  sudo apt-get install poppler-utils
    Windows: Download poppler and add to PATH

Run:
    streamlit run underwriter_app.py
"""

import streamlit as st
import anthropic
import pandas as pd
import sqlite3
import json
import base64
import io
import re
import os
import math
import copy
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

# Try importing pdf2image — guide user if missing
try:
    from pdf2image import convert_from_bytes
    PDF2IMAGE_AVAILABLE = True
except ImportError:
    PDF2IMAGE_AVAILABLE = False

# Try importing openpyxl utilities for Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ---------------------------------------------------------------------------
# BRANDING & COLORS (from Wellx brand assets)
# ---------------------------------------------------------------------------
WELLX_ORANGE = "#fb9b35"
WELLX_PINK = "#f1517b"
WELLX_MAGENTA = "#b43082"
WELLX_NAVY = "#003780"
WELLX_SKY = "#35c5fc"
WELLX_VIOLET = "#8431cb"

# ---------------------------------------------------------------------------
# API KEY — loaded from Streamlit secrets (for cloud) or fallback
# Set in .streamlit/secrets.toml locally, or in Streamlit Cloud dashboard.
# ---------------------------------------------------------------------------
ANTHROPIC_API_KEY = st.secrets.get("ANTHROPIC_API_KEY", "")

# ---------------------------------------------------------------------------
# PLAN & COMMISSION DEFAULTS
# ---------------------------------------------------------------------------
PLAN_OPTIONS = ["HealthX-QIC", "OpenX"]

# Commission structures per plan
COMMISSION_DEFAULTS = {
    "HealthX-QIC": {
        "Broker": 10.0,
        "HealthX": 3.0,
        "QIC": 0.0,
        "NAS": 4.0,
        "Insurance Tax": 0.5,
        "Reinsurance Margin": 7.0,
    },
    "OpenX": {
        "Broker": 15.0,
        "OpenX": 3.0,
        "DNI": 3.0,
        "NAS": 4.0,
        "RI Broker": 1.5,
        "Insurance Tax": 0.5,
        "Reinsurance Margin": 7.0,
    },
}

# Relation types that count as dependents
DEPENDENT_RELATIONS = {"spouse", "wife", "husband", "son", "daughter", "child"}
# Relation types that count as principals/employees
PRINCIPAL_RELATIONS = {"principal", "employee", "self"}
# Allowed relation types (principals + dependents)
ALLOWED_RELATIONS = PRINCIPAL_RELATIONS | DEPENDENT_RELATIONS | {"others", "other"}

# ---------------------------------------------------------------------------
# DATABASE SETUP — SQLite for persisting quotes
# ---------------------------------------------------------------------------
DB_PATH = Path(__file__).parent / "quotes.db"


def get_db():
    """Return a connection to the SQLite database, creating tables if needed."""
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    conn.execute("""
        CREATE TABLE IF NOT EXISTS quotes (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at    TEXT    NOT NULL,
            company_name  TEXT    NOT NULL,
            broker_name   TEXT,
            status        TEXT    DEFAULT 'neutral',
            summary_json  TEXT,
            raw_extract   TEXT,
            commission_broker  REAL,
            commission_insurer REAL,
            commission_tpa     REAL,
            commission_wellx   REAL,
            commission_margins REAL,
            burning_cost       REAL,
            indicative_premium REAL,
            current_census     INTEGER,
            notes         TEXT
        )
    """)
    conn.commit()
    return conn


def save_quote(data: dict):
    """Insert a new quote record and return its ID."""
    conn = get_db()
    cur = conn.execute("""
        INSERT INTO quotes (
            created_at, company_name, broker_name, status,
            summary_json, raw_extract,
            commission_broker, commission_insurer, commission_tpa,
            commission_wellx, commission_margins,
            burning_cost, indicative_premium, current_census, notes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        datetime.now().isoformat(),
        data.get("company_name", ""),
        data.get("broker_name", ""),
        data.get("status", "neutral"),
        json.dumps(data.get("summary", {})),
        json.dumps(data.get("raw_extract", {})),
        data.get("commission_broker", 10),
        data.get("commission_insurer", 0.5),
        data.get("commission_tpa", 4),
        data.get("commission_wellx", 4),
        data.get("commission_margins", 7),
        data.get("burning_cost", 0),
        data.get("indicative_premium", 0),
        data.get("current_census", 0),
        data.get("notes", ""),
    ))
    conn.commit()
    quote_id = cur.lastrowid
    conn.close()
    return quote_id


def update_quote_status(quote_id: int, status: str):
    """Update the status of an existing quote."""
    conn = get_db()
    conn.execute("UPDATE quotes SET status = ? WHERE id = ?", (status, quote_id))
    conn.commit()
    conn.close()


def get_all_quotes():
    """Return all quotes as a list of dicts."""
    conn = get_db()
    rows = conn.execute("SELECT * FROM quotes ORDER BY created_at DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_quote_by_id(quote_id: int):
    """Return a single quote by ID."""
    conn = get_db()
    row = conn.execute("SELECT * FROM quotes WHERE id = ?", (quote_id,)).fetchone()
    conn.close()
    return dict(row) if row else None


# ---------------------------------------------------------------------------
# CLAUDE VISION EXTRACTION — the core AI parsing engine
# ---------------------------------------------------------------------------

# This system prompt is the backbone of extraction accuracy.  It tells Claude
# exactly which sections of a DHA report to look for, what data to pull, and
# how to structure the JSON response.  It is intentionally very detailed so
# that Claude can handle scanned, rotated, or non-standard DHA PDFs.

EXTRACTION_SYSTEM_PROMPT = """You are an expert Medical Insurance Underwriter AI specialized in parsing DHA (Dubai Health Authority) insurance reports and custom claims reports.

Your task is to extract ALL relevant data from the provided PDF pages and return a single, well-structured JSON object.

=== DHA REPORT SECTIONS TO EXTRACT ===

SECTION 1 — Scheme/Employer Info
- employer_name: The name of the scheme or employer (exact text)

SECTION 3 — Policy Period
- policy_effective_date: dd-mm-yy or dd/mm/yyyy format
- policy_expiry_date: dd-mm-yy or dd/mm/yyyy format
- initial_policy_effective_date: the date continuous cover started

SECTION 4 — Report Period
- report_period_start: dd-mm-yy or dd/mm/yyyy
- report_period_end: dd-mm-yy or dd/mm/yyyy
- report_production_date: dd-mm-yy or dd/mm/yyyy

SECTION 5 — Claims Values (UAE Dirham)
- claims_paid: Value of claims processed during the report period (number)
- claims_outstanding: Value of claims incurred, reported but not processed (number)
- claims_ibnr: Value of claims incurred but not reported (number)

SECTION 6 — Population Census (Beginning of reporting period)
Return a dictionary with keys: male, single_female, married_female
Each maps to age buckets: {"0_15": n, "16_25": n, "26_35": n, "36_50": n, "51_65": n, "over_65": n, "total": n}
Also provide "grand_total" summing all categories.

SECTION 7 — Population Census (End of reporting period)
Same structure as Section 6 but for the ending census.

SECTION 8 — Claims by Member Type (UAE Dirham)
Return rows for: employee, spouse, dependents, totals
Each row has: ip, op, pharmacy, dental, optical, total

SECTION 10 — Claims by Diagnosis (Top 10 by value)
Return a list of objects: [{diagnosis: str, ip: number, op: number, total: number}, ...]

SECTION 11 — Number of Claims by Diagnosis (corresponds to Section 10)
Return a list of objects: [{diagnosis: str, ip: number, op: number, total: number}, ...]

SECTION 12 — Claims by Provider (Top 10)
Return a list of objects: [{provider: str, ip: number, op: number, total: number}, ...]

SECTION 13 — Number of Claims by Provider
Return a list of objects: [{provider: str, ip: number, op: number, total: number}, ...]

SECTION 17 — Monthly Claims
Return a list of objects in order: [{month: str, year: number, value: number}, ...]
Include ALL 12 months (17a through 17l). Use 0 for months with no data.
Label months clearly: e.g. "MAY", "JUNE", etc.

SECTION 19 — Complex Cases (CRITICAL — often unlabelled, look just below Section 18)
Look for any section after Section 18 that lists complex or ongoing cases.
It may be titled "COMPLEX CASES", "ONGOING CASES", or have no title at all.
Extract any text found here as a string. Look for conditions like:
Neoplasms, Polycythemia Vera, Myelodysplastic Syndromes, Cancer, Chemotherapy, etc.
Return as: "complex_cases_notes": "full text of section 19 or empty string if not found"

=== IMPORTANT PARSING RULES ===

1. Numbers: Remove commas, convert to numeric. "3,095,350" → 3095350
2. Dates: Keep original format but also provide ISO (yyyy-mm-dd) versions where possible.
3. Missing data: Use null for genuinely missing values, 0 for confirmed zero values.
4. Section numbers: Look for bold or large section numbers (1, 3, 4, 5, 6, 7, 8, 10, 11, 12, 13, 17).
5. If a table spans multiple pages, combine the data.
6. If the document is NOT a standard DHA report (e.g., MaxHealth Client Performance Analytics, custom claims report), extract whatever structured claims data you can find and note "non_standard_format": true.
   For MaxHealth reports specifically:
   - employer_name: from the title page (e.g. "Arab Broadcasting Services FZ-LLC")
   - policy dates: use the "Incurred Period" dates as policy_effective_date and policy_expiry_date
   - claims_paid: "Total Paid Claims" or the Totals row from "Total Paid Claims" table
   - claims_outstanding: "Outstanding Claims" value from Policy Overview
   - claims_ibnr: Total Incurred minus (Paid + Outstanding), or null if not derivable
   - claims_by_member_type: from "Claims by Member Type" tables (Employee/Spouse/Dependents)
   - diagnosis_top10: from "Top 10 Diagnosis Categories" table
   - provider_top10: from "Top 10 Providers" table
   - monthly_claims: from "Claims Spend by Month" chart/table
   - complex_cases_notes: from "Top 5 Claimants" section (member conditions)
   - census data: extract "total_members" (the Total Members count) and "membership_change_pct" (e.g. -2.4 if "decrease of 2.4%"). Put these as top-level fields. Also fill census_start/census_end if you can derive them.
7. For monthly claims (Section 17), months with empty/blank values should be set to 0.
8. Always look for a "Total" row or column and capture it.

=== OUTPUT FORMAT ===

Return ONLY a valid JSON object with this top-level structure:
{
  "employer_name": "...",
  "policy_effective_date": "...",
  "policy_expiry_date": "...",
  "initial_policy_effective_date": "...",
  "report_period_start": "...",
  "report_period_end": "...",
  "report_production_date": "...",
  "claims_paid": 0,
  "claims_outstanding": 0,
  "claims_ibnr": 0,
  "census_start": { ... },
  "census_end": { ... },
  "claims_by_member_type": { ... },
  "diagnosis_top10_values": [ ... ],
  "diagnosis_top10_counts": [ ... ],
  "provider_top10_values": [ ... ],
  "provider_top10_counts": [ ... ],
  "monthly_claims": [ ... ],
  "complex_cases_notes": "",
  "non_standard_format": false,
  "extraction_notes": "Any notes about data quality, missing sections, etc."
}

Do NOT include any text outside the JSON object. Do NOT wrap in markdown code fences.
Return ONLY the raw JSON."""


def extract_dha_report_with_claude(api_key: str, pdf_bytes: bytes) -> dict:
    """
    Convert PDF pages to images, send to Claude Vision, and extract structured data.

    Steps:
      1. Use pdf2image to convert each page to a PNG image.
      2. Base64-encode all page images.
      3. Send all images to Claude 3.5 Sonnet with the extraction prompt.
      4. Parse the JSON response.
    """
    if not PDF2IMAGE_AVAILABLE:
        st.error(
            "pdf2image is not installed. Run: `pip install pdf2image`\n"
            "Also install poppler:\n"
            "  macOS: `brew install poppler`\n"
            "  Ubuntu: `sudo apt-get install poppler-utils`"
        )
        return {}

    # Convert PDF pages to images (200 DPI is a good balance of quality vs size)
    with st.spinner("Converting PDF pages to images..."):
        try:
            images = convert_from_bytes(pdf_bytes, dpi=200, fmt="png")
        except Exception as e:
            st.error(f"Failed to convert PDF to images: {e}")
            return {}

    st.info(f"Converted {len(images)} page(s). Sending to Jasper for Analysis...")

    # Build the message content: one image per page + the extraction request
    content_blocks = []
    for i, img in enumerate(images):
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        b64 = base64.standard_b64encode(buf.getvalue()).decode("utf-8")
        content_blocks.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": "image/png",
                "data": b64,
            },
        })
        content_blocks.append({
            "type": "text",
            "text": f"[Page {i + 1} of {len(images)}]",
        })

    content_blocks.append({
        "type": "text",
        "text": (
            "Please analyze ALL the pages above. They form a single insurance report "
            "(it may be a DHA report, MaxHealth Client Performance Analytics, or another format). "
            "Extract every piece of data according to your instructions and return the JSON."
        ),
    })

    # Call Claude
    client = anthropic.Anthropic(api_key=api_key)

    with st.spinner("Jasper is analyzing the report — this may take 30-60 seconds..."):
        try:
            response = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=8000,
                system=EXTRACTION_SYSTEM_PROMPT,
                messages=[{"role": "user", "content": content_blocks}],
            )
        except anthropic.AuthenticationError:
            st.error("Invalid API key. Please check your Anthropic API key.")
            return {}
        except Exception as e:
            st.error(f"Claude API error: {e}")
            return {}

    # Parse the response
    raw_text = response.content[0].text.strip()

    # Strip markdown fences if Claude included them despite instructions
    if raw_text.startswith("```"):
        raw_text = re.sub(r"^```(?:json)?\s*", "", raw_text)
        raw_text = re.sub(r"\s*```$", "", raw_text)

    try:
        data = json.loads(raw_text)
    except json.JSONDecodeError:
        st.error("Claude returned non-JSON output. Showing raw response for debugging.")
        st.code(raw_text[:3000])
        return {}

    return data


# ---------------------------------------------------------------------------
# SOP CALCULATIONS
# ---------------------------------------------------------------------------

def parse_date_flexible(date_str: str) -> Optional[datetime]:
    """Try multiple date formats to parse a date string."""
    if not date_str or date_str in ("null", "None", "00-Jan-00"):
        return None
    formats = [
        "%d-%m-%y", "%d-%m-%Y", "%d/%m/%y", "%d/%m/%Y",
        "%Y-%m-%d", "%m/%d/%Y", "%d-%b-%y", "%d-%b-%Y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None


def consolidate_census_files(files) -> tuple:
    """
    Accept a list of uploaded Streamlit file objects (Excel/CSV), detect their
    format (Liva/Nextcare 93-col vs MaxHealth/Chrome 41-col vs standard census),
    normalize to a common schema, filter to active members, and return a single
    consolidated DataFrame plus a summary dict.

    Returns (consolidated_df, summary_dict).
    summary_dict keys: per_file (list of {name, format, total, active, inactive}),
                       total_active, total_inactive
    """
    # ── Column mappings for each known format ──
    LIVA_HEADER_MARKER = "Beneficiary FullName"      # 93-col Nextcare/Liva
    MAX_HEADER_MARKER  = "INSURED NAME"              # 41-col MaxHealth/Chrome

    STANDARD_COLS = [
        "Source File", "Category", "Beneficiary Name",
        "DOB", "Age", "Gender", "Nationality",
        "Relation", "Marital Status", "Emirates ID", "Passport No",
        "Card Number", "Policy No", "Annual Limit", "Emirate", "Status",
    ]

    all_rows = []
    summary_per_file = []

    for f in files:
        fname = f.name
        ext = fname.split(".")[-1].lower()

        # Read into a raw DataFrame, auto-detecting the header row.
        # Liva/Nextcare files often have a title row + blank row before the
        # real header, so if the default read doesn't surface known columns
        # we scan the first 10 rows for the actual header.
        KNOWN_HEADERS = {
            "beneficiary fullname", "insured name", "member name",
            "payer", "master_contract", "relation", "gender",
        }
        try:
            f.seek(0)
            if ext == "csv":
                raw_df = pd.read_csv(f)
            else:
                raw_df = pd.read_excel(f)

            # Check if known headers are in columns
            cur_cols = {str(c).strip().lower() for c in raw_df.columns}
            if not (cur_cols & KNOWN_HEADERS):
                # Scan first rows for the real header
                for scan_row in range(min(10, len(raw_df))):
                    row_vals = {str(v).strip().lower() for v in raw_df.iloc[scan_row] if pd.notna(v)}
                    if row_vals & KNOWN_HEADERS:
                        # Re-read with this row as header
                        f.seek(0)
                        if ext == "csv":
                            raw_df = pd.read_csv(f, header=scan_row + 1)
                        else:
                            raw_df = pd.read_excel(f, header=scan_row + 1)
                        break
        except Exception as exc:
            summary_per_file.append({"name": fname, "format": "ERROR", "total": 0,
                                     "active": 0, "inactive": 0, "error": str(exc)})
            continue

        # Drop fully blank rows
        raw_df = raw_df.dropna(how="all").reset_index(drop=True)
        col_lower = {c.strip().lower(): c for c in raw_df.columns}
        col_set = set(col_lower.keys())

        detected_format = "standard"
        file_rows = []

        # ── Detect Liva / Nextcare format ─────────────────────────────────
        if "beneficiary fullname" in col_set or ("card number" in col_set and "category" in col_set and raw_df.shape[1] > 60):
            detected_format = "Liva/Nextcare"
            c = col_lower  # shorthand

            def _g(row, key):
                col = c.get(key)
                if col is None:
                    return ""
                v = row.get(col)
                return str(v).strip() if pd.notna(v) else ""

            for _, row in raw_df.iterrows():
                status = _g(row, "status")
                is_active = status.lower() == "active" if status else True

                # Parse DOB
                dob = ""
                age = _g(row, "age")
                dob_col = c.get("dob")
                if dob_col and pd.notna(row.get(dob_col)):
                    try:
                        dob = str(row[dob_col])[:10]
                    except Exception:
                        pass

                file_rows.append({
                    "Source File": fname,
                    "Category": _g(row, "category"),
                    "Beneficiary Name": _g(row, "beneficiary fullname"),
                    "DOB": dob,
                    "Age": age,
                    "Gender": _g(row, "gender"),
                    "Nationality": _g(row, "nationality"),
                    "Relation": _g(row, "dependency"),
                    "Marital Status": _g(row, "marital status"),
                    "Emirates ID": _g(row, "national identityno"),
                    "Passport No": _g(row, "passportno"),
                    "Card Number": _g(row, "card number"),
                    "Policy No": _g(row, "policy.no"),
                    "Annual Limit": _g(row, "annual limit"),
                    "Emirate": _g(row, "emirate-visaissued") or _g(row, "person work.emirate"),
                    "Status": status,
                    "_active": is_active,
                })

        # ── Detect MaxHealth / Chrome format ──────────────────────────────
        elif "insured name" in col_set or ("member name" in col_set and "plcy_ref" in col_set):
            detected_format = "MaxHealth/Chrome"
            c = col_lower

            def _g(row, key):
                col = c.get(key)
                if col is None:
                    return ""
                v = row.get(col)
                return str(v).strip() if pd.notna(v) else ""

            for _, row in raw_df.iterrows():
                status = _g(row, "status")
                is_active = status.lower() == "active" if status else True

                dob = ""
                dob_col = c.get("date of birth")
                if dob_col and pd.notna(row.get(dob_col)):
                    try:
                        dob = str(row[dob_col])[:10]
                    except Exception:
                        pass

                age = ""
                if dob:
                    try:
                        bd = datetime.strptime(dob, "%Y-%m-%d")
                        today = datetime.today()
                        age = str(today.year - bd.year - ((today.month, today.day) < (bd.month, bd.day)))
                    except Exception:
                        pass

                # Normalize relation
                relation = _g(row, "relation to principal")
                emirate = _g(row, "visa_issued")

                file_rows.append({
                    "Source File": fname,
                    "Category": _g(row, "category"),
                    "Beneficiary Name": _g(row, "insured name") or _g(row, "member name"),
                    "DOB": dob,
                    "Age": age,
                    "Gender": _g(row, "gender"),
                    "Nationality": _g(row, "nationality"),
                    "Relation": relation,
                    "Marital Status": _g(row, "marital status"),
                    "Emirates ID": _g(row, "emirates id"),
                    "Passport No": _g(row, "passport_no"),
                    "Card Number": _g(row, "member id"),
                    "Policy No": _g(row, "plcy_ref"),
                    "Annual Limit": _g(row, "sum assured"),
                    "Emirate": emirate,
                    "Status": status,
                    "_active": is_active,
                })

        # ── Standard / simple census format ───────────────────────────────
        else:
            detected_format = "Standard"
            # Try to map common columns
            c = col_lower

            def _find_key(candidates):
                # Exact match first, then prefix match so headers like
                # "Relationship (Principal, Spouse, Child)" or
                # "Date of birth (dd/mm/yyyy)" still resolve to the right column.
                for cand in candidates:
                    if cand in c:
                        return cand
                for cand in candidates:
                    for actual in c:
                        if actual.startswith(cand):
                            return actual
                return None

            rel_key = _find_key(("relation", "relationship", "member type", "dependency", "relation to principal"))
            gender_key = _find_key(("gender", "sex"))
            dob_key = _find_key(("date of birth", "dob", "birth date", "birthdate", "date_of_birth"))
            name_key = _find_key(("beneficiary fullname", "insured name", "member name", "name", "full name", "employee name"))
            nationality_key = _find_key(("nationality",))
            status_key = _find_key(("status",))
            eid_key = _find_key(("emirates id", "national identityno", "eid"))

            def _g(row, key):
                if key is None:
                    return ""
                col = c.get(key)
                if col is None:
                    return ""
                v = row.get(col)
                return str(v).strip() if pd.notna(v) else ""

            for _, row in raw_df.iterrows():
                status = _g(row, status_key) if status_key else ""
                is_active = status.lower() == "active" if status else True  # no status col → assume active

                dob = ""
                if dob_key:
                    dob_col_name = c.get(dob_key)
                    if dob_col_name and pd.notna(row.get(dob_col_name)):
                        try:
                            dob = str(row[dob_col_name])[:10]
                        except Exception:
                            pass

                age = ""
                if dob:
                    try:
                        bd = datetime.strptime(dob, "%Y-%m-%d")
                        today = datetime.today()
                        age = str(today.year - bd.year - ((today.month, today.day) < (bd.month, bd.day)))
                    except Exception:
                        pass

                file_rows.append({
                    "Source File": fname,
                    "Category": "",
                    "Beneficiary Name": _g(row, name_key) if name_key else "",
                    "DOB": dob,
                    "Age": age,
                    "Gender": _g(row, gender_key) if gender_key else "",
                    "Nationality": _g(row, nationality_key) if nationality_key else "",
                    "Relation": _g(row, rel_key) if rel_key else "",
                    "Marital Status": "",
                    "Emirates ID": _g(row, eid_key) if eid_key else "",
                    "Passport No": "",
                    "Card Number": "",
                    "Policy No": "",
                    "Annual Limit": "",
                    "Emirate": "",
                    "Status": status,
                    "_active": is_active,
                })

        # ── Tally active/inactive for this file ──
        active_rows = [r for r in file_rows if r["_active"]]
        inactive_count = len(file_rows) - len(active_rows)
        summary_per_file.append({
            "name": fname,
            "format": detected_format,
            "total": len(file_rows),
            "active": len(active_rows),
            "inactive": inactive_count,
        })
        all_rows.extend(active_rows)

    # ── Build consolidated DataFrame ──
    if not all_rows:
        return pd.DataFrame(columns=STANDARD_COLS), {
            "per_file": summary_per_file, "total_active": 0, "total_inactive": 0
        }

    consolidated_df = pd.DataFrame(all_rows)[STANDARD_COLS]

    # ── Normalize Relation column for downstream analyze_census_file ──
    relation_map = {
        "principal": "Principal", "employee": "Principal", "self": "Principal",
        "spouse": "Spouse", "wife": "Spouse", "husband": "Spouse",
        "child": "Child", "son": "Child", "daughter": "Child",
        "others": "Others",
    }
    consolidated_df["Relation"] = (
        consolidated_df["Relation"]
        .str.strip().str.lower()
        .map(relation_map)
        .fillna(consolidated_df["Relation"])
    )

    total_inactive = sum(f["inactive"] for f in summary_per_file)
    summary = {
        "per_file": summary_per_file,
        "total_active": len(consolidated_df),
        "total_inactive": total_inactive,
    }
    return consolidated_df, summary


def analyze_census_file(df: pd.DataFrame) -> dict:
    """
    Analyze a census DataFrame (from Excel/CSV upload) and return structured
    breakdown: employee vs dependent counts, age distribution, married female
    flags, and any disallowed relation types.

    Expects columns: RELATION (or Relation), GENDER (or Gender),
    Date Of Birth (or DOB / DATE OF BIRTH).
    """
    result = {
        "total_members": len(df),
        "employees": 0,
        "dependents": 0,
        "employee_pct": 0.0,
        "dependent_pct": 0.0,
        "married_females_18_45": 0,
        "married_females_18_45_pct": 0.0,
        "age_distribution": {},
        "flags": [],
        "disallowed_relations": [],
    }

    # Drop fully blank rows before counting
    df = df.dropna(how="all").reset_index(drop=True)
    total = len(df)
    result["total_members"] = total
    if total == 0:
        return result

    # ── Normalize column names to lowercase for flexible matching ──
    col_map = {c.strip().lower(): c for c in df.columns}

    # Find relation column
    rel_col = None
    for key in ("relation", "relationship", "member type"):
        if key in col_map:
            rel_col = col_map[key]
            break

    # Find gender column
    gender_col = None
    for key in ("gender", "sex"):
        if key in col_map:
            gender_col = col_map[key]
            break

    # Find DOB column
    dob_col = None
    for key in ("date of birth", "dob", "birth date", "birthdate", "date_of_birth"):
        if key in col_map:
            dob_col = col_map[key]
            break

    # ── Employee / Dependent split ──
    if rel_col:
        relations = df[rel_col].astype(str).str.strip().str.lower()

        employees = relations.isin(PRINCIPAL_RELATIONS).sum()
        dependents = relations.isin(DEPENDENT_RELATIONS).sum()

        result["employees"] = int(employees)
        result["dependents"] = int(dependents)
        result["employee_pct"] = round(employees / total * 100, 1)
        result["dependent_pct"] = round(dependents / total * 100, 1)

        # ── Flag disallowed relation types ──
        unique_relations = set(relations.unique())
        disallowed = unique_relations - ALLOWED_RELATIONS - {""}
        if disallowed:
            result["disallowed_relations"] = sorted(disallowed)
            result["flags"].append(
                f"DISALLOWED RELATION TYPES: {', '.join(sorted(disallowed))}. "
                "Only Principal/Employee, Spouse, Wife, Son, Daughter, Child are allowed."
            )

    # ── Age distribution ──
    ages = None
    if dob_col:
        try:
            dob_series = pd.to_datetime(df[dob_col], errors="coerce")
            today = pd.Timestamp.now()
            ages = ((today - dob_series).dt.days / 365.25).dropna().astype(int)
        except Exception:
            pass

    if ages is not None and len(ages) > 0:
        bins = [0, 10, 20, 30, 40, 50, 60, 70, 200]
        labels = ["0-10", "11-20", "21-30", "31-40", "41-50", "51-60", "61-70", "70+"]
        age_groups = pd.cut(ages, bins=bins, labels=labels, right=True)
        dist = age_groups.value_counts().sort_index()
        for label in labels:
            count = int(dist.get(label, 0))
            pct = round(count / total * 100, 1)
            result["age_distribution"][label] = {"count": count, "pct": pct}

        # ── Married Females 18-45 ──
        if rel_col and gender_col:
            is_married_female = (
                relations.isin({"spouse", "wife"}) &
                df[gender_col].astype(str).str.strip().str.upper().str.startswith("F")
            )
            in_age_range = (ages >= 18) & (ages <= 45)
            # Align indices
            combined = is_married_female & in_age_range.reindex(is_married_female.index, fill_value=False)
            mf_count = int(combined.sum())
            mf_pct = round(mf_count / total * 100, 1)
            result["married_females_18_45"] = mf_count
            result["married_females_18_45_pct"] = mf_pct

            if mf_pct > 15:
                result["flags"].append(
                    f"HIGH MARRIED FEMALES (18-45): {mf_count} ({mf_pct}%) — exceeds 15% threshold."
                )

        # ── Flag if >10% are aged 50-99 ──
        aged_50_plus = int(((ages >= 50) & (ages < 100)).sum())
        aged_50_pct = round(aged_50_plus / total * 100, 1)
        result["aged_50_plus"] = aged_50_plus
        result["aged_50_plus_pct"] = aged_50_pct
        if aged_50_pct > 10:
            result["flags"].append(
                f"HIGH SENIOR RATIO (50+): {aged_50_plus} ({aged_50_pct}%) — exceeds 10% threshold."
            )

    return result


def run_sop_analysis(data: dict, commissions: dict, company_name: str, plan: str = "HealthX-QIC", uploaded_census_count: int = 0) -> dict:
    """
    Run the full DHA SOP analysis on extracted data.

    Returns a summary dict with all calculated values and flags.
    """
    summary = {
        "company_name": company_name,
        "validations": {},
        "claims_analysis": {},
        "census_analysis": {},
        "diagnosis_analysis": {},
        "burning_cost_analysis": {},
        "premium_quotation": {},
        "flags": [],
    }

    # -----------------------------------------------------------------------
    # STEP 1: VALIDATION & RECENCY
    # -----------------------------------------------------------------------
    employer_name = data.get("employer_name", "")
    summary["employer_name_extracted"] = employer_name

    # Employer match check
    if company_name.strip().lower() in employer_name.strip().lower() or \
       employer_name.strip().lower() in company_name.strip().lower():
        summary["validations"]["employer_match"] = True
    else:
        summary["validations"]["employer_match"] = False
        summary["flags"].append(
            f"EMPLOYER MISMATCH: Input '{company_name}' vs Report '{employer_name}'"
        )

    # Policy effective date validity (< 365 days from today)
    policy_eff = parse_date_flexible(data.get("policy_effective_date", ""))
    policy_exp = parse_date_flexible(data.get("policy_expiry_date", ""))
    today = datetime.now()

    if policy_eff:
        days_since = (today - policy_eff).days
        summary["validations"]["policy_effective_date"] = data.get("policy_effective_date")
        summary["validations"]["policy_days_since"] = days_since
        summary["validations"]["policy_valid"] = days_since < 365
        if days_since >= 365:
            summary["flags"].append("POLICY EXPIRED: Effective date is more than 365 days ago.")
    else:
        summary["validations"]["policy_valid"] = None
        summary["flags"].append("Could not parse policy effective date.")

    if policy_exp:
        summary["validations"]["policy_expiry_date"] = data.get("policy_expiry_date")

    # Report recency (end date not older than 90 days)
    report_end = parse_date_flexible(data.get("report_period_end", ""))
    if report_end:
        days_old = (today - report_end).days
        summary["validations"]["report_end_date"] = data.get("report_period_end")
        summary["validations"]["report_days_old"] = days_old
        summary["validations"]["report_recent"] = days_old <= 90
        if days_old > 90:
            summary["flags"].append(
                f"STALE REPORT: Report end date is {days_old} days old (>90 days)."
            )
    else:
        summary["validations"]["report_recent"] = None

    # -----------------------------------------------------------------------
    # STEP 2: CLAIMS RATIO & MEMBER TYPE
    # -----------------------------------------------------------------------
    claims_paid = float(data.get("claims_paid", 0) or 0)
    claims_outstanding = float(data.get("claims_outstanding", 0) or 0)
    claims_ibnr = float(data.get("claims_ibnr", 0) or 0)
    total_incurred = claims_paid + claims_outstanding + claims_ibnr

    summary["claims_analysis"]["claims_paid"] = claims_paid
    summary["claims_analysis"]["claims_outstanding"] = claims_outstanding
    summary["claims_analysis"]["claims_ibnr"] = claims_ibnr
    summary["claims_analysis"]["total_incurred"] = total_incurred

    # Outstanding ratio
    if claims_paid > 0:
        outstanding_ratio = (claims_outstanding / claims_paid) * 100
    else:
        outstanding_ratio = 0
    summary["claims_analysis"]["outstanding_ratio_pct"] = round(outstanding_ratio, 2)
    summary["claims_analysis"]["outstanding_benchmark"] = ">20%"
    if outstanding_ratio > 20:
        summary["flags"].append(
            f"HIGH OUTSTANDING RATIO: {outstanding_ratio:.1f}% (benchmark >20%)"
        )

    # IP Ratio from Section 8
    member_type = data.get("claims_by_member_type", {})
    totals_row = member_type.get("totals", member_type.get("Totals", {}))
    ip_total = float(totals_row.get("ip", 0) or 0)
    claims_total_s8 = float(totals_row.get("total", 0) or 0)
    if claims_total_s8 > 0:
        ip_ratio = (ip_total / claims_total_s8) * 100
    else:
        ip_ratio = 0
    summary["claims_analysis"]["ip_total"] = ip_total
    summary["claims_analysis"]["claims_total_section8"] = claims_total_s8
    summary["claims_analysis"]["ip_ratio_pct"] = round(ip_ratio, 2)
    summary["claims_analysis"]["ip_benchmark"] = "20-25%"

    # -----------------------------------------------------------------------
    # STEP 3: CENSUS & POPULATION GROWTH
    # -----------------------------------------------------------------------
    def sum_census(census_data):
        """Sum all values in a census section."""
        if not census_data:
            return 0
        total = census_data.get("grand_total", 0)
        if total:
            return int(total)
        # Fallback: sum across categories
        s = 0
        for cat in ("male", "single_female", "married_female"):
            cat_data = census_data.get(cat, {})
            if isinstance(cat_data, dict):
                cat_total = cat_data.get("total", 0)
                if cat_total:
                    s += int(cat_total)
                else:
                    for k, v in cat_data.items():
                        if k != "total":
                            s += int(v or 0)
            elif isinstance(cat_data, (int, float)):
                s += int(cat_data)
        return s

    census_start = sum_census(data.get("census_start"))
    census_end = sum_census(data.get("census_end"))

    # Census fallback for non-standard reports (MaxHealth etc.)
    census_assumed = False
    if census_start == 0 and census_end == 0:
        total_members = data.get("total_members", 0)
        change_pct_val = data.get("membership_change_pct", None)
        if total_members and total_members > 0:
            if change_pct_val is not None:
                census_start = int(total_members)
                census_end = int(round(total_members * (1 + change_pct_val / 100)))
            else:
                census_end = int(total_members)
        # Fallback: assume starting = uploaded census * 0.95
        uploaded_census = 0
        census_analysis_data = st.session_state.get("last_census_analysis")
        if census_analysis_data:
            uploaded_census = census_analysis_data.get("total_members", 0)
        if census_start == 0 and census_end == 0 and uploaded_census > 0:
            census_start = int(round(uploaded_census * 0.95))
            census_assumed = True
        if census_start > 0 and census_end == 0 and uploaded_census > 0:
            start_str = data.get("policy_effective_date") or data.get("report_period_start", "")
            end_str = data.get("report_period_end", "")
            expiry_str = data.get("policy_expiry_date", "")
            rp_start = parse_date_flexible(start_str)
            rp_end = parse_date_flexible(end_str)
            pol_expiry = parse_date_flexible(expiry_str)
            if rp_start and rp_end and pol_expiry and pol_expiry > rp_start:
                total_days = (pol_expiry - rp_start).days
                elapsed_days = (rp_end - rp_start).days
                if total_days > 0:
                    census_end = int(round(
                        census_start + (uploaded_census - census_start) * elapsed_days / total_days
                    ))
                else:
                    census_end = uploaded_census
            else:
                census_end = uploaded_census

    avg_census = (census_start + census_end) / 2 if (census_start + census_end) > 0 else 1

    if census_start > 0:
        census_change_pct = ((census_end - census_start) / census_start) * 100
    else:
        census_change_pct = 0

    summary["census_analysis"]["census_start"] = census_start
    summary["census_analysis"]["census_end"] = census_end
    summary["census_analysis"]["census_assumed"] = census_assumed
    summary["census_analysis"]["avg_census"] = avg_census
    summary["census_analysis"]["census_change_pct"] = round(census_change_pct, 2)
    summary["census_analysis"]["benchmark"] = "±15%"

    if abs(census_change_pct) > 15:
        summary["flags"].append(
            f"CENSUS CHANGE: {census_change_pct:+.1f}% (outside ±15% benchmark). "
            "Request monthly census from broker."
        )

    # -----------------------------------------------------------------------
    # STEP 4: DIAGNOSIS & PROVIDER EVALUATION
    # -----------------------------------------------------------------------
    diag_values = data.get("diagnosis_top10_values", [])
    diag_counts = data.get("diagnosis_top10_counts", [])

    major_conditions_keywords = [
        "neoplasm", "cancer", "chemotherapy", "autoimmune", "malignant",
        "carcinoma", "tumor", "tumour", "lymphoma", "leukemia", "leukaemia",
    ]
    flagged_conditions = []
    high_value_claims = []

    for i, diag in enumerate(diag_values):
        name = diag.get("diagnosis", "").lower()
        total_val = float(diag.get("total", 0) or 0)

        # Check for major conditions
        for kw in major_conditions_keywords:
            if kw in name:
                flagged_conditions.append(diag.get("diagnosis", ""))
                break

        # Per-claim value (Section 10 / Section 11)
        if i < len(diag_counts):
            count_total = float(diag_counts[i].get("total", 0) or 0)
            if count_total > 0:
                per_claim = total_val / count_total
                if per_claim > 30000:
                    high_value_claims.append({
                        "diagnosis": diag.get("diagnosis", ""),
                        "per_claim_aed": round(per_claim, 2),
                        "total_value": total_val,
                        "claim_count": count_total,
                    })

    # Top-10 concentration
    top10_sum = sum(float(d.get("total", 0) or 0) for d in diag_values)
    if claims_paid > 0:
        top10_concentration = (top10_sum / claims_paid) * 100
    else:
        top10_concentration = 0

    summary["diagnosis_analysis"]["flagged_conditions"] = flagged_conditions
    summary["diagnosis_analysis"]["high_value_claims"] = high_value_claims
    summary["diagnosis_analysis"]["top10_sum"] = top10_sum
    summary["diagnosis_analysis"]["top10_concentration_pct"] = round(top10_concentration, 2)

    if flagged_conditions:
        summary["flags"].append(
            f"MAJOR CONDITIONS DETECTED: {', '.join(flagged_conditions)}"
        )
    if high_value_claims:
        names = [h["diagnosis"] for h in high_value_claims]
        summary["flags"].append(
            f"HIGH VALUE CLAIMS (>30k AED/claim): {', '.join(names)}"
        )

    # -----------------------------------------------------------------------
    # STEP 5: BURNING COST CALCULATION
    # -----------------------------------------------------------------------
    monthly = data.get("monthly_claims", [])
    monthly_values = [float(m.get("value", 0) or 0) for m in monthly]

    # Determine which months have actual data (non-zero)
    incurred_months = [v for v in monthly_values if v > 0]
    n_incurred = len(incurred_months)

    # Determine policy start day to decide averaging method
    policy_start_day = 1  # default
    if policy_eff:
        policy_start_day = policy_eff.day

    # Sum check
    monthly_sum = sum(monthly_values)
    sum_check_vs_paid = abs(monthly_sum - claims_paid) if claims_paid else 0
    summary["burning_cost_analysis"]["monthly_sum"] = monthly_sum
    summary["burning_cost_analysis"]["sum_check_vs_paid"] = round(sum_check_vs_paid, 2)
    summary["burning_cost_analysis"]["sum_matches_paid"] = sum_check_vs_paid < (claims_paid * 0.05) if claims_paid else True

    # Build the three averages based on policy start day
    if n_incurred >= 3:
        if policy_start_day <= 5:
            # Method for 1st-5th: use all incurred months
            avg_a = sum(incurred_months) / len(incurred_months)
            avg_b = sum(incurred_months[:-1]) / len(incurred_months[:-1]) if len(incurred_months) > 1 else avg_a
            avg_c = sum(incurred_months[:-2]) / len(incurred_months[:-2]) if len(incurred_months) > 2 else avg_b
            method = "1st-5th (include first month)"
        else:
            # Method for 6th+: exclude first month
            excl_first = incurred_months[1:]
            avg_a = sum(excl_first) / len(excl_first) if excl_first else 0
            excl_first_last = incurred_months[1:-1]
            avg_b = sum(excl_first_last) / len(excl_first_last) if excl_first_last else avg_a
            excl_first_last2 = incurred_months[1:-2]
            avg_c = sum(excl_first_last2) / len(excl_first_last2) if excl_first_last2 else avg_b
            method = "6th+ (exclude first month)"
    elif n_incurred > 0:
        avg_a = sum(incurred_months) / len(incurred_months)
        avg_b = avg_a
        avg_c = avg_a
        method = "Insufficient months — using simple average"
    else:
        avg_a = avg_b = avg_c = 0
        method = "No incurred months"

    highest_avg = max(avg_a, avg_b, avg_c)

    summary["burning_cost_analysis"]["policy_start_day"] = policy_start_day
    summary["burning_cost_analysis"]["method"] = method
    summary["burning_cost_analysis"]["n_incurred_months"] = n_incurred
    summary["burning_cost_analysis"]["avg_a"] = round(avg_a, 2)
    summary["burning_cost_analysis"]["avg_b"] = round(avg_b, 2)
    summary["burning_cost_analysis"]["avg_c"] = round(avg_c, 2)
    summary["burning_cost_analysis"]["highest_avg_monthly"] = round(highest_avg, 2)

    # Burning Cost per member per month
    if avg_census > 0:
        burning_cost_per_capita = highest_avg / avg_census
    else:
        burning_cost_per_capita = 0

    summary["burning_cost_analysis"]["burning_cost_per_capita"] = round(burning_cost_per_capita, 2)

    # Adjustments
    inflation = 0.05  # +5%
    ip_adjustment = 0
    if ip_ratio < 20:
        ip_adjustment = (25 - ip_ratio) / 100  # Bring up to 25%

    outstanding_adjustment = 0
    if outstanding_ratio > 20:
        outstanding_adjustment = (outstanding_ratio - 20) / 100

    adjusted_burning_cost = burning_cost_per_capita * (1 + inflation + ip_adjustment + outstanding_adjustment)

    summary["burning_cost_analysis"]["inflation_pct"] = inflation * 100
    summary["burning_cost_analysis"]["ip_adjustment_pct"] = round(ip_adjustment * 100, 2)
    summary["burning_cost_analysis"]["outstanding_adjustment_pct"] = round(outstanding_adjustment * 100, 2)
    summary["burning_cost_analysis"]["adjusted_burning_cost_per_capita"] = round(adjusted_burning_cost, 2)

    # -----------------------------------------------------------------------
    # STEP 6: PREMIUM QUOTATION
    # -----------------------------------------------------------------------
    # Prefer uploaded census count over DHA report census
    if uploaded_census_count > 0:
        current_census = uploaded_census_count
    else:
        current_census = census_end if census_end > 0 else census_start
    projected_claims = adjusted_burning_cost * 12 * current_census

    # Net premium (before commissions) — same as projected claims unless manual adjustments
    net_premium = projected_claims

    # Sum all commission items
    total_commission_pct_val = sum(commissions.values())
    total_commission_pct = total_commission_pct_val / 100

    # Total excluding Reinsurance Margin
    ri_margin = commissions.get("Reinsurance Margin", commissions.get("margins", 7))
    total_excl_ri = total_commission_pct_val - ri_margin

    if total_commission_pct < 1:
        indicative_premium = net_premium / (1 - total_commission_pct)
    else:
        indicative_premium = net_premium
        summary["flags"].append("WARNING: Total commission >= 100%, cannot calculate premium.")

    # Determine the platform key (HealthX or OpenX) for actual commission calc
    platform_key = None
    platform_pct = 0
    for key in commissions:
        if key.lower() in ("healthx", "openx"):
            platform_key = key
            platform_pct = commissions[key]
            break

    # Platform actual commission = Indicative Premium * platform %
    actual_commission = indicative_premium * (platform_pct / 100) if platform_pct > 0 else 0

    # Minimum platform % = (110 * census) / Indicative Premium
    min_platform_pct = (110 * current_census) / indicative_premium if indicative_premium > 0 else 0

    summary["premium_quotation"]["plan"] = plan
    summary["premium_quotation"]["current_census"] = current_census
    summary["premium_quotation"]["projected_claims_annual"] = round(projected_claims, 2)
    summary["premium_quotation"]["net_premium"] = round(net_premium, 2)
    summary["premium_quotation"]["total_commission_pct"] = round(total_commission_pct * 100, 2)
    summary["premium_quotation"]["total_excl_ri_margin_pct"] = round(total_excl_ri, 2)
    summary["premium_quotation"]["indicative_premium"] = round(indicative_premium, 2)
    summary["premium_quotation"]["premium_per_member_annual"] = (
        round(indicative_premium / current_census, 2) if current_census > 0 else 0
    )
    summary["premium_quotation"]["premium_per_member_monthly"] = (
        round(indicative_premium / current_census / 12, 2) if current_census > 0 else 0
    )
    summary["premium_quotation"]["platform_key"] = platform_key or plan.split("-")[0]
    summary["premium_quotation"]["platform_pct"] = platform_pct
    summary["premium_quotation"]["actual_commission"] = round(actual_commission, 2)
    summary["premium_quotation"]["min_platform_pct"] = round(min_platform_pct, 4)

    return summary


# ---------------------------------------------------------------------------
# EXCEL EXPORT — 3 sheets, styled to match Claims_Analysis reference template
# ---------------------------------------------------------------------------

# Template colours (matching reference Claims_Analysis_*.xlsx)
XLS_NAVY = "003780"          # title bar + value-cell header text
XLS_LABEL_FILL = "EEF3F8"    # left-column label cells
XLS_SUBTITLE_CYAN = "35C5FC" # "from the Intake Form..." subtitle
XLS_CURRENCY_FMT = '#,##0'
XLS_PCT_FMT = '0.00%'
XLS_DATE_FMT = '[$]dd\\ mmm\\ yyyy;@'


def _parse_date(value):
    """Parse a date string into a datetime; returns None on failure."""
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d %b %Y", "%d %B %Y",
                "%Y/%m/%d", "%b %d, %Y", "%B %d, %Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def generate_quote_excel(
    summary: dict,
    data: dict,
    commissions: dict,
    prepared_by: str = "",
    broker_name: str = "",
    rm_name: str = "",
) -> bytes:
    """
    Build a 3-sheet workbook matching the Claims_Analysis_*.xlsx reference:
      1. Prospect Info-Extracted
      2. BC & Premium Calculation
      3. Summary
    All cross-sheet references and totals are live formulas so underwriters
    can tweak inputs and see recalculated outputs. Returns xlsx bytes.
    """
    wb = Workbook()

    # --- shared styles / helpers ---------------------------------------
    thin = Side(style="thin")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    label_fill = PatternFill(start_color=XLS_LABEL_FILL, end_color=XLS_LABEL_FILL, fill_type="solid")
    navy_fill = PatternFill(start_color=XLS_NAVY, end_color=XLS_NAVY, fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    label_font = Font(name="Calibri", size=12, color=XLS_NAVY)
    value_font = Font(name="Calibri", size=12, color="000000")
    section_font = Font(name="Calibri", bold=True, size=12, color=XLS_NAVY)
    section_fill = PatternFill(start_color=XLS_LABEL_FILL, end_color=XLS_LABEL_FILL, fill_type="solid")

    def write_label(ws, coord, text):
        c = ws[coord]
        c.value = text
        c.font = label_font
        c.fill = label_fill
        c.border = thin_border

    def write_value(ws, coord, value, number_format=None, bold=False, font_color="000000"):
        c = ws[coord]
        c.value = value
        c.font = Font(name="Calibri", size=12, color=font_color, bold=bold)
        c.fill = white_fill
        c.border = thin_border
        if number_format:
            c.number_format = number_format
        return c

    def write_note(ws, coord, text):
        """Column-C explanatory note — grey italic-ish small text like reference."""
        c = ws[coord]
        c.value = text
        c.font = Font(name="Calibri", size=11, color="5E788A")

    def write_section_header(ws, coord, text):
        c = ws[coord]
        c.value = text
        c.font = section_font
        c.fill = section_fill

    # =======================================================================
    # SHEET 1: Prospect Info-Extracted
    # =======================================================================
    ws1 = wb.active
    ws1.title = "Prospect Info-Extracted"

    # Column widths matching reference
    widths1 = {"A": 72, "B": 20, "C": 23, "D": 9, "E": 20, "F": 11}
    for letter, w in widths1.items():
        ws1.column_dimensions[letter].width = w

    # --- Title bar (rows 1-2) ---
    ws1.merge_cells("A1:C1")
    t1 = ws1["A1"]
    t1.value = "Prospect Information"
    t1.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    t1.fill = navy_fill
    t1.alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[1].height = 24

    sub = ws1["A2"]
    sub.value = "from the Intake Form and Extracted information"
    sub.font = Font(name="Calibri", size=12, color=XLS_SUBTITLE_CYAN)
    sub.fill = navy_fill
    # Fill rest of row 2 navy for banded look
    for col in ("B2", "C2"):
        ws1[col].fill = navy_fill

    # --- Intake / extracted info (rows 4-10) ---
    company_name = summary.get("company_name", "") or ""
    employer_extracted = summary.get("employer_name_extracted", "") or ""
    employer_match = (summary.get("validations") or {}).get("employer_match", False)
    plan_value = (summary.get("premium_quotation") or {}).get("plan", "") or ""

    expiry_raw = (summary.get("validations") or {}).get("policy_expiry_date", "")
    expiry_dt = _parse_date(expiry_raw)

    write_label(ws1, "A4", "Company Name")
    write_value(ws1, "B4", company_name)

    write_label(ws1, "A5", "Company Name  (from Report)")
    write_value(ws1, "B5", employer_extracted)
    if employer_match:
        write_note(ws1, "C5", "Company Names matching")
    elif employer_extracted:
        write_note(ws1, "C5", "Company Names NOT matching")

    write_label(ws1, "A6", "Expiry Date")
    if expiry_dt:
        c = write_value(ws1, "B6", expiry_dt, number_format=XLS_DATE_FMT)
        c.alignment = Alignment(horizontal="center")
    else:
        write_value(ws1, "B6", expiry_raw or "")

    write_label(ws1, "A8", "Broker")
    write_value(ws1, "B8", broker_name)

    write_label(ws1, "A9", "Relationship Manager")
    write_value(ws1, "B9", rm_name)

    write_label(ws1, "A10", "Plan")
    write_value(ws1, "B10", plan_value)

    # --- CLAIMS OVERVIEW (rows 12-16) ---
    write_section_header(ws1, "A12", "CLAIMS OVERVIEW")

    ca = summary.get("claims_analysis", {}) or {}
    write_label(ws1, "A13", "Claims Paid (AED)")
    write_value(ws1, "B13", float(ca.get("claims_paid", 0) or 0), number_format=XLS_CURRENCY_FMT)

    write_label(ws1, "A14", "Claims Outstanding (AED)")
    write_value(ws1, "B14", float(ca.get("claims_outstanding", 0) or 0), number_format=XLS_CURRENCY_FMT)
    cc14 = ws1["C14"]
    cc14.value = "=B14/$B$13"
    cc14.number_format = XLS_PCT_FMT
    cc14.font = Font(name="Calibri", size=11, color="5E788A")

    write_label(ws1, "A15", "Claims IBNR (AED)")
    write_value(ws1, "B15", float(ca.get("claims_ibnr", 0) or 0), number_format=XLS_CURRENCY_FMT)

    write_label(ws1, "A16", "Total Incurred (AED)")
    c16 = write_value(ws1, "B16", "=SUM(B13:B15)", number_format=XLS_CURRENCY_FMT, bold=True)
    ws1["A16"].font = Font(name="Calibri", bold=True, size=12, color=XLS_NAVY)

    # --- CLAIMS BY TYPE (rows 20-26) ---
    write_section_header(ws1, "A20", "CLAIMS BY TYPE")

    totals = ((data.get("claims_by_member_type") or {}).get("totals")) or {}
    type_rows = [
        ("IP", totals.get("ip", 0)),
        ("OP", totals.get("op", 0)),
        ("Pharmacy", totals.get("pharmacy", 0)),
        ("Dental", totals.get("dental", 0)),
        ("Optical", totals.get("optical", 0)),
    ]
    for i, (lbl, val) in enumerate(type_rows):
        r = 21 + i
        write_label(ws1, f"A{r}", lbl)
        write_value(ws1, f"B{r}", float(val or 0), number_format=XLS_CURRENCY_FMT)
        cc = ws1[f"C{r}"]
        cc.value = f"=B{r}/$B$26"
        cc.number_format = XLS_PCT_FMT
        cc.font = Font(name="Calibri", size=11, color="5E788A")

    # Row 26: sum of claims-by-type
    sum_cell = ws1["B26"]
    sum_cell.value = "=SUM(B21:B25)"
    sum_cell.number_format = XLS_CURRENCY_FMT
    sum_cell.font = Font(name="Calibri", bold=True, size=12, color="000000")
    sum_cell.border = thin_border
    write_note(ws1, "C26", "Matching with claims paid")

    # --- CENSUS INFORMATION (rows 29-34) ---
    write_section_header(ws1, "A29", "CENSUS INFORMATION")

    census = summary.get("census_analysis", {}) or {}
    write_label(ws1, "A30", "Starting Census")
    write_value(ws1, "B30", int(census.get("census_start", 0) or 0))

    write_label(ws1, "A31", "Ending Census")
    write_value(ws1, "B31", int(census.get("census_end", 0) or 0))

    write_label(ws1, "A33", "Average Census")
    write_value(ws1, "B33", "=AVERAGE(B30:B31)")

    write_label(ws1, "A34", "Census Change")
    write_value(ws1, "B34", "=B31/B30-1", number_format=XLS_PCT_FMT)

    # --- TOP 10 DIAGNOSES / PROVIDERS (rows 37-39) ---
    diag_values = data.get("diagnosis_top10_values", []) or []
    prov_values = data.get("provider_top10_values", []) or []
    diag_total = sum(float(d.get("total", 0) or 0) for d in diag_values)
    prov_total = sum(float(p.get("total", 0) or 0) for p in prov_values)

    write_label(ws1, "A37", "TOP 10 DIAGNOSES TOTAL")
    write_value(ws1, "B37", diag_total, number_format=XLS_CURRENCY_FMT)
    cc = ws1["C37"]
    cc.value = "=B37/$B$13"
    cc.number_format = XLS_PCT_FMT
    cc.font = Font(name="Calibri", size=11, color="5E788A")

    write_label(ws1, "A39", "TOP 10 PROVIDERS TOTAL")
    write_value(ws1, "B39", prov_total, number_format=XLS_CURRENCY_FMT)
    cc = ws1["C39"]
    cc.value = "=B39/$B$13"
    cc.number_format = XLS_PCT_FMT
    cc.font = Font(name="Calibri", size=11, color="5E788A")

    # --- Monthly Claims (rows 42-53) ---
    write_section_header(ws1, "A42", "Monthly Claims")

    monthly = data.get("monthly_claims", []) or []
    # Keep up to 10 rows; row 43..52 match the reference's fixed block
    for i in range(10):
        r = 43 + i
        if i < len(monthly):
            m = monthly[i]
            month_lbl = f"{m.get('month', '')} {m.get('year', '')}".strip()
            write_label(ws1, f"A{r}", month_lbl)
            write_value(
                ws1, f"B{r}",
                float(m.get("value", 0) or 0),
                number_format=XLS_CURRENCY_FMT,
            )
        else:
            write_label(ws1, f"A{r}", "")
            write_value(ws1, f"B{r}", None, number_format=XLS_CURRENCY_FMT)

    sum53 = ws1["B53"]
    sum53.value = "=SUM(B43:B52)"
    sum53.number_format = XLS_CURRENCY_FMT
    sum53.font = Font(name="Calibri", bold=True, size=12, color="000000")
    sum53.border = thin_border
    write_note(ws1, "C53", "Matching with claims paid")

    # --- COMPLEX CASES (row 55+) ---
    write_section_header(ws1, "A55", "COMPLEX CASES")
    flagged = (summary.get("diagnosis_analysis") or {}).get("flagged_conditions", []) or []
    next_row = 56
    if flagged:
        for i, item in enumerate(flagged):
            r = 56 + i
            c = ws1[f"A{r}"]
            c.value = item
            c.font = Font(name="Calibri", size=12, color="000000")
            c.alignment = Alignment(wrap_text=True, vertical="top")
            next_row = r + 1
    else:
        ws1["A56"].value = "None flagged"
        ws1["A56"].font = Font(name="Calibri", size=12, color="5E788A")
        next_row = 57

    # Prepared-by footer (two rows below last complex-case)
    now = datetime.now()
    prep_text = f"Prepared by: {prepared_by or '—'} on {now.strftime('%d %b %Y')} at {now.strftime('%I:%M %p').lstrip('0')}"
    footer = ws1[f"A{next_row + 1}"]
    footer.value = prep_text
    footer.font = Font(name="Calibri", italic=True, size=11, color="5E788A")

    # =======================================================================
    # SHEET 2: BC & Premium Calculation
    # =======================================================================
    ws2 = wb.create_sheet("BC & Premium Calculation")

    widths2 = {"A": 30, "B": 18, "C": 18, "D": 28, "E": 22}
    for letter, w in widths2.items():
        ws2.column_dimensions[letter].width = w

    # Header row (row 1) — navy fill, white text, centred
    headers = ["Monthly Claims", "Paid Claims", "Claims Haircut", "Selected Claims after Haircut", "Selected"]
    for idx, h in enumerate(headers):
        c = ws2.cell(row=1, column=idx + 1, value=h)
        c.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        c.fill = navy_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    # Monthly rows (2..11)
    monthly_included = st.session_state.get("monthly_included", []) if hasattr(st, "session_state") else []
    for i in range(10):
        r = 2 + i
        if i < len(monthly):
            m = monthly[i]
            ws2.cell(row=r, column=1, value=f"{m.get('month', '')} {m.get('year', '')}".strip())
            ws2.cell(row=r, column=2, value=float(m.get("value", 0) or 0)).number_format = XLS_CURRENCY_FMT
            ws2.cell(row=r, column=3, value=0).number_format = XLS_CURRENCY_FMT
            # Determine default selected state: use session flag if available, else first/last excluded
            if monthly_included and i < len(monthly_included):
                selected = bool(monthly_included[i])
            else:
                selected = not (i == 0 or i == len(monthly) - 1)
            # D is live-linked to E: toggling the Selected checkbox recalculates the average.
            d = ws2.cell(row=r, column=4, value=f'=IF(E{r},B{r}-C{r},"")')
            d.number_format = XLS_CURRENCY_FMT
            ws2.cell(row=r, column=5, value=selected)
        else:
            ws2.cell(row=r, column=5, value=False)

        for col in range(1, 6):
            ws2.cell(row=r, column=col).border = thin_border

    # Row 12: sum of Paid Claims
    s12 = ws2.cell(row=12, column=2, value="=SUM(B2:B11)")
    s12.number_format = XLS_CURRENCY_FMT
    s12.font = Font(name="Calibri", bold=True, size=12)
    s12.border = thin_border

    # --- BURNING COST CALCULATION block ---
    bc = summary.get("burning_cost_analysis", {}) or {}

    hdr14 = ws2["A14"]
    hdr14.value = "BURNING COST CALCULATION"
    hdr14.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    hdr14.fill = navy_fill
    hdr14.alignment = Alignment(horizontal="left", vertical="center")
    for col in ("B14", "C14"):
        ws2[col].fill = navy_fill

    # Rows 15-34
    bc_rows = [
        ("A15", "Policy Start Day", "B15", int(bc.get("policy_start_day", 1) or 1), None, None),
        ("A16", "Incurred Months", "B16", int(bc.get("n_incurred_months", 0) or 0), None, None),
        ("A18", "Monthly Average", "B18", "=AVERAGE(D2:D11)", XLS_CURRENCY_FMT, None),
        ("A19", "Average Census", "B19", "='Prospect Info-Extracted'!B33", None, None),
        ("A20", "Monthly Burning Cost Per Capita", "B20", "=B18/B19", XLS_CURRENCY_FMT, None),
        ("A22", "Inflation Adjustment", "B22", float(bc.get("inflation_pct", 0) or 0) / 100, XLS_PCT_FMT, None),
        ("A23", "IP Adjustment", "B23", float(bc.get("ip_adjustment_pct", 0) or 0) / 100, XLS_PCT_FMT, None),
        ("A24", "Outstanding Adjustment", "B24", float(bc.get("outstanding_adjustment_pct", 0) or 0) / 100, XLS_PCT_FMT, None),
        ("A25", "Adjusted Monthly Burning Cost Per Capita ", "B25", "=B20*(1+SUM(B22:B24))", XLS_CURRENCY_FMT, None),
        ("A27", "Yearly Burning Cost per Capita", "B27", "=B25*12", XLS_CURRENCY_FMT, "Expected annual claims per member"),
        ("A28", "Current Census", "B28",
            int((summary.get("premium_quotation") or {}).get("current_census", 0) or 0), None, None),
        ("A30", "Over reporting Haircut (%)", "B30", 0.10, XLS_PCT_FMT, "i.e. (assumed over reported claims)"),
        ("A31", "Major Claims Allowance", "B31", 150000, XLS_CURRENCY_FMT, None),
        ("A34", "Yearly Burning Cost", "B34", "=(B27*B28)*(1-B30)+B31", XLS_CURRENCY_FMT, "Expected annual claims"),
    ]
    for a_coord, label, b_coord, value, nf, note in bc_rows:
        write_label(ws2, a_coord, label)
        c = write_value(ws2, b_coord, value, number_format=nf)
        if note:
            c_coord = "C" + b_coord[1:]
            write_note(ws2, c_coord, note)

    # --- INDICATIVE PREMIUM block ---
    hdr36 = ws2["A36"]
    hdr36.value = "INDICATIVE PREMIUM"
    hdr36.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    hdr36.fill = navy_fill
    for col in ("B36", "C36"):
        ws2[col].fill = navy_fill

    write_label(ws2, "A39", "INDICATIVE PREMIUM")
    ws2["A39"].font = Font(name="Calibri", bold=True, size=12, color=XLS_NAVY)
    ws2["A39"].fill = label_fill
    ws2["A39"].border = thin_border
    ip_cell = write_value(ws2, "B39", "=B34/(1-B47)", number_format=XLS_CURRENCY_FMT, bold=True, font_color=XLS_NAVY)

    # Row 40 — average premium helper
    avg_cell = write_value(ws2, "B40", "=B39/B28", number_format=XLS_CURRENCY_FMT)
    write_note(ws2, "C40", "average premium")

    # Commission rows (42-46) — populated from commissions dict based on plan
    plan = (summary.get("premium_quotation") or {}).get("plan", "")

    # Choose ordered commission keys that match the reference layout:
    # Row 42 = Broker, 43 = platform secondary (QIC/DNI), 44 = Healthx/OpenX (platform),
    # 45 = NAS, 46 = Reinsurance Margin (+ insurance tax/RI broker folded in if present)
    comm_ordered: list[tuple[str, float]] = []
    broker_label = f"Broker ({broker_name})" if broker_name else "Broker"
    broker_pct = float(commissions.get("Broker", commissions.get("broker", 0)) or 0)
    comm_ordered.append((broker_label, broker_pct))

    if "QIC" in commissions:
        comm_ordered.append(("QIC", float(commissions.get("QIC", 0) or 0)))
    elif "DNI" in commissions:
        comm_ordered.append(("DNI", float(commissions.get("DNI", 0) or 0)))
    else:
        comm_ordered.append(("Insurance Tax", float(commissions.get("Insurance Tax", 0) or 0)))

    if "HealthX" in commissions:
        platform_label, platform_pct = "Healthx", float(commissions.get("HealthX", 0) or 0)
    elif "OpenX" in commissions:
        platform_label, platform_pct = "OpenX", float(commissions.get("OpenX", 0) or 0)
    else:
        platform_label, platform_pct = "Healthx", 0.0
    comm_ordered.append((platform_label, platform_pct))

    comm_ordered.append(("Nas", float(commissions.get("NAS", commissions.get("Nas", 0)) or 0)))

    # Row 46: bundle remaining (RI Broker + Insurance Tax + Reinsurance Margin)
    ri_margin = float(commissions.get("Reinsurance Margin", 0) or 0)
    # If plan has RI Broker + Insurance Tax as separate, roll them into row 46 label
    extras = 0.0
    extra_labels = []
    if plan and "OpenX" in plan:
        extras += float(commissions.get("RI Broker", 0) or 0)
        extras += float(commissions.get("Insurance Tax", 0) or 0)
        if commissions.get("RI Broker", 0):
            extra_labels.append("RI Broker")
        if commissions.get("Insurance Tax", 0):
            extra_labels.append("Ins. Tax")
    ri_label = "Reinsurance Margin" if not extra_labels else f"Reinsurance Margin (+{', '.join(extra_labels)})"
    comm_ordered.append((ri_label, ri_margin + extras))

    # Write rows 42..46
    for j, (lbl, pct) in enumerate(comm_ordered[:5]):
        r = 42 + j
        write_label(ws2, f"A{r}", lbl)
        write_value(ws2, f"B{r}", float(pct) / 100.0, number_format=XLS_PCT_FMT)
        cc = ws2[f"C{r}"]
        cc.value = f"=B{r}*$B$39"
        cc.number_format = XLS_CURRENCY_FMT
        cc.font = value_font
        cc.border = thin_border

    # D44 / E44: per-member AED + minimum note (reference row 44)
    d44 = ws2["D44"]
    d44.value = "=C44/$B$28"
    d44.number_format = XLS_CURRENCY_FMT
    d44.font = value_font
    d44.border = thin_border

    e44 = ws2["E44"]
    e44.value = "AED 110 (USD 30) minimum"
    e44.font = Font(name="Calibri", size=11, color="5E788A")

    # Row 47: total commission
    t47 = ws2["B47"]
    t47.value = "=SUM(B42:B46)"
    t47.number_format = XLS_PCT_FMT
    t47.font = Font(name="Calibri", bold=True, size=12)
    t47.border = thin_border

    # =======================================================================
    # SHEET 3: Summary
    # =======================================================================
    ws3 = wb.create_sheet("Summary")
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 32

    title3 = ws3["A1"]
    title3.value = "Executive Summary"
    title3.font = Font(name="Calibri", bold=True, size=14, color=XLS_NAVY)

    write_label(ws3, "A3", "Company Name")
    write_value(ws3, "B3", company_name)

    write_label(ws3, "A4", "Expiry Date")
    if expiry_dt:
        ce = write_value(ws3, "B4", expiry_dt, number_format=XLS_DATE_FMT)
        ce.alignment = Alignment(horizontal="center")
    else:
        write_value(ws3, "B4", expiry_raw or "")

    write_label(ws3, "A6", "Broker")
    write_value(ws3, "B6", broker_name)

    write_label(ws3, "A7", "Relationship Manager")
    write_value(ws3, "B7", rm_name)

    # Email draft (rows 9-10, merged, wrap text)
    eh = ws3["A9"]
    eh.value = "Email Draft"
    eh.font = Font(name="Calibri", bold=True, size=12, color=XLS_NAVY)
    eh.fill = label_fill

    expiry_str = expiry_dt.strftime("%d %b %Y") if expiry_dt else (expiry_raw or "TBD")
    email_body = (
        f"Hi {rm_name or 'Team'}, \n\n"
        f"Please see below INDICATIVE premium for {company_name or 'the prospect'} "
        f"whose policy is expiring on {expiry_str}, thru {broker_name or 'the broker'}:"
    )
    ws3.merge_cells("A10:C10")
    em = ws3["A10"]
    em.value = email_body
    em.font = Font(name="Calibri", size=11, color="000000")
    em.alignment = Alignment(wrap_text=True, vertical="top")
    ws3.row_dimensions[10].height = 60

    # Premium figures (rows 13-15) — cross-sheet references to sheet 2
    write_label(ws3, "A13", "INDICATIVE PREMIUM")
    ws3["A13"].font = Font(name="Calibri", bold=True, size=12, color=XLS_NAVY)
    write_value(
        ws3, "B13",
        "='BC & Premium Calculation'!B39",
        number_format=XLS_CURRENCY_FMT, bold=True, font_color=XLS_NAVY,
    )

    write_label(ws3, "A14", "Census")
    write_value(ws3, "B14", "='BC & Premium Calculation'!B19")

    write_label(ws3, "A15", "Per Capita")
    write_value(ws3, "B15", "=B13/B14", number_format=XLS_CURRENCY_FMT)

    # AI findings (rows 17-18) — merged, wrap-text
    fh = ws3["A17"]
    fh.value = "AI Underwriter Findings & Additional requirements"
    fh.font = Font(name="Calibri", bold=True, size=12, color=XLS_NAVY)
    fh.fill = label_fill

    flags = summary.get("flags", []) or []
    if flags:
        findings_text = "\n".join(f"• {f}" for f in flags)
    else:
        findings_text = (
            "{Prompt: Mention all the flags here. i.e. Ask for recent report if report "
            "is more than 90 days, ask for monthly census if more than 10%, ask for recent "
            "medical report and lab results for ongoing major conditions, et al.}"
        )
    ws3.merge_cells("A18:C18")
    ff = ws3["A18"]
    ff.value = findings_text
    ff.font = Font(name="Calibri", size=11, color="000000")
    ff.alignment = Alignment(wrap_text=True, vertical="top")
    # Dynamic row height based on number of lines
    ws3.row_dimensions[18].height = max(60, min(30 + 15 * (findings_text.count("\n") + 1), 320))

    # --- Save to bytes ---
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ---------------------------------------------------------------------------
# STREAMLIT UI
# ---------------------------------------------------------------------------

def setup_page():
    """Configure page layout and custom CSS adapted from the WellX Premium Summary Tool."""
    st.set_page_config(
        page_title="WellX AI Underwriter",
        page_icon="🏥",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    # ---------------------------------------------------------------------------
    # Inject Google Fonts + full WellX-inspired CSS
    # Adapted from templates/index.html — uses the same brand palette but with
    # a teal accent (#0d9488) to distinguish the Underwriter from the Summary Tool.
    # ---------------------------------------------------------------------------
    st.markdown("""
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Raleway:wght@700;800;900&display=swap" rel="stylesheet">

    <style>
    /* ── CSS Variables (WellX palette + teal accent) ───────────────────────── */
    :root {
        --orange:  #fb9b35;
        --pink:    #f1517b;
        --magenta: #b43082;
        --navy:    #003780;
        --sky:     #35c5fc;
        --violet:  #8431cb;
        --slate:   #5e788a;
        --dark:    #0a0a0a;
        --teal:    #0d9488;
        --grad-full: linear-gradient(90deg, #fb9b35, #f1517b, #b43082, #8431cb, #35c5fc);
        --grad-warm: linear-gradient(135deg, #fb9b35, #f1517b, #b43082);
        --grad-cool: linear-gradient(135deg, #003780, #8431cb, #35c5fc);
        --grad-teal: linear-gradient(135deg, #0d9488, #003780, #8431cb);
    }

    /* ── Global overrides ──────────────────────────────────────────────────── */
    .stApp {
        font-family: 'Inter', system-ui, sans-serif !important;
    }
    .stApp > header {
        background: var(--grad-full) !important;
        height: 4px !important;
    }

    /* ── Sidebar — dark brand panel ────────────────────────────────────────── */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0a0a0a 0%, #111827 100%) !important;
    }
    [data-testid="stSidebar"] * {
        color: #ffffff !important;
    }
    [data-testid="stSidebar"] .stRadio label span {
        font-family: 'Raleway', sans-serif !important;
        font-weight: 700 !important;
        font-size: 0.92rem !important;
        letter-spacing: 0.3px;
    }
    [data-testid="stSidebar"] hr {
        border-color: rgba(255,255,255,0.1) !important;
    }

    /* ── Section labels (Raleway uppercase like index.html) ────────────────── */
    .section-lbl {
        font-family: 'Raleway', sans-serif;
        font-size: 0.7rem;
        font-weight: 800;
        letter-spacing: 2px;
        text-transform: uppercase;
        color: var(--navy);
        margin: 28px 0 14px;
        padding-bottom: 9px;
        border-bottom: 2px solid transparent;
        border-image: var(--grad-full) 1;
    }
    .section-lbl:first-child { margin-top: 0; }

    /* ── Stat cards (mirrors .stat-card from index.html) ───────────────────── */
    .stat-card {
        background: #f8faff;
        border: 1.5px solid #e0e7f0;
        border-radius: 11px;
        padding: 16px 18px;
        text-align: center;
        transition: 0.15s;
    }
    .stat-card:hover {
        border-color: var(--navy);
        box-shadow: 0 4px 18px rgba(0,55,128,0.08);
    }
    .stat-card .stat-val {
        font-family: 'Raleway', sans-serif;
        font-weight: 900;
        font-size: 1.15rem;
        color: var(--navy);
        margin-bottom: 4px;
    }
    .stat-card .stat-lbl {
        font-size: 0.72rem;
        font-weight: 600;
        color: var(--slate);
        text-transform: uppercase;
        letter-spacing: 0.8px;
    }
    .stat-card .stat-sub {
        font-size: 0.7rem;
        color: var(--slate);
        margin-top: 3px;
    }
    .stat-card.highlight {
        border-color: var(--orange);
        background: #fff9f0;
    }
    .stat-card.highlight .stat-val {
        color: var(--orange);
    }
    .stat-card.teal {
        border-color: var(--teal);
        background: linear-gradient(135deg, #f0fdfa, #f8faff);
    }
    .stat-card.teal .stat-val {
        color: var(--teal);
    }

    /* ── Premium hero card ─────────────────────────────────────────────────── */
    .premium-hero {
        background: linear-gradient(135deg, #003780 0%, #8431cb 50%, #35c5fc 100%);
        border-radius: 14px;
        padding: 24px 28px;
        color: #fff;
        text-align: center;
        box-shadow: 0 8px 32px rgba(0,55,128,0.3);
    }
    .premium-hero .ph-label {
        font-family: 'Raleway', sans-serif;
        font-size: 0.72rem;
        font-weight: 800;
        letter-spacing: 2px;
        text-transform: uppercase;
        opacity: 0.8;
        margin-bottom: 6px;
    }
    .premium-hero .ph-value {
        font-family: 'Raleway', sans-serif;
        font-size: 1.8rem;
        font-weight: 900;
        margin-bottom: 6px;
    }
    .premium-hero .ph-sub {
        font-size: 0.78rem;
        opacity: 0.75;
    }

    /* ── Validation badges (like recon-banner in index.html) ───────────────── */
    .badge-ok {
        background: linear-gradient(135deg, #d4edda, #c3e6cb);
        border-radius: 10px;
        padding: 12px 16px;
        margin-bottom: 8px;
        display: flex;
        align-items: center;
        gap: 10px;
        font-size: 0.88rem;
        font-weight: 600;
        color: #155724;
    }
    .badge-fail {
        background: linear-gradient(135deg, #f8d7da, #f5c6cb);
        border-radius: 10px;
        padding: 12px 16px;
        margin-bottom: 8px;
        display: flex;
        align-items: center;
        gap: 10px;
        font-size: 0.88rem;
        font-weight: 600;
        color: #721c24;
    }
    .badge-warn {
        background: linear-gradient(135deg, #fff3cd, #ffe9a0);
        border-radius: 10px;
        padding: 12px 16px;
        margin-bottom: 8px;
        display: flex;
        align-items: center;
        gap: 10px;
        font-size: 0.88rem;
        font-weight: 600;
        color: #856404;
    }

    /* ── Info box ───────────────────────────────────────────────────────────── */
    .info-box {
        background: #f0f4ff;
        border: 1.5px solid #c8d8f8;
        border-radius: 9px;
        padding: 12px 16px;
        font-size: 0.82rem;
        color: #374151;
        line-height: 1.55;
        margin-bottom: 18px;
    }
    .info-box strong { color: var(--navy); }

    /* ── Commission table (mirrors .commission-table) ──────────────────────── */
    .comm-tbl {
        width: 100%;
        border-collapse: collapse;
        border-radius: 11px;
        overflow: hidden;
        border: 1.5px solid #e0e7f0;
        margin-bottom: 16px;
    }
    .comm-tbl thead tr {
        background: linear-gradient(90deg, rgba(0,55,128,0.05), rgba(53,197,252,0.05));
    }
    .comm-tbl thead th {
        padding: 10px 16px;
        font-family: 'Raleway', sans-serif;
        font-size: 0.68rem;
        font-weight: 800;
        text-transform: uppercase;
        letter-spacing: 1.2px;
        color: var(--navy);
        text-align: left;
    }
    .comm-tbl tbody tr {
        border-top: 1px solid #eef1f6;
    }
    .comm-tbl tbody tr:hover {
        background: #f8f9fd;
    }
    .comm-tbl td {
        padding: 10px 16px;
        font-size: 0.88rem;
    }
    .comm-tbl tfoot tr {
        background: #f0f4ff;
        border-top: 2px solid var(--navy);
    }
    .comm-tbl tfoot td {
        font-weight: 700;
        color: var(--navy);
        font-family: 'Raleway', sans-serif;
        padding: 10px 16px;
    }

    /* ── Buttons — match grad-cool / grad-warm look ────────────────────────── */
    .stButton > button[kind="primary"],
    .stDownloadButton > button {
        background: var(--grad-teal) !important;
        color: #fff !important;
        border: none !important;
        border-radius: 10px !important;
        font-family: 'Raleway', sans-serif !important;
        font-weight: 800 !important;
        letter-spacing: 0.5px !important;
        box-shadow: 0 4px 18px rgba(13,148,136,0.3) !important;
        transition: 0.15s !important;
    }
    .stButton > button[kind="primary"]:hover,
    .stDownloadButton > button:hover {
        opacity: 0.9 !important;
        box-shadow: 0 6px 24px rgba(13,148,136,0.4) !important;
    }

    /* ── File uploader zone ─────────────────────────────────────────────────── */
    [data-testid="stFileUploader"] {
        border: 2px dashed #c8d4e3 !important;
        border-radius: 11px !important;
        transition: 0.15s;
    }
    [data-testid="stFileUploader"]:hover {
        border-color: var(--navy) !important;
        background: rgba(0,55,128,0.02) !important;
    }

    /* ── Input fields — match .field input style ───────────────────────────── */
    .stTextInput input, .stNumberInput input, .stSelectbox select {
        border: 1.5px solid #d1d9e6 !important;
        border-radius: 9px !important;
        font-family: 'Inter', sans-serif !important;
        transition: 0.15s !important;
    }
    .stTextInput input:focus, .stNumberInput input:focus {
        border-color: var(--navy) !important;
        box-shadow: 0 0 0 3px rgba(0,55,128,0.1) !important;
    }

    /* ── Expander styling ──────────────────────────────────────────────────── */
    .streamlit-expanderHeader {
        font-family: 'Raleway', sans-serif !important;
        font-weight: 700 !important;
        color: var(--navy) !important;
    }

    /* ── Dataframe / table overrides ───────────────────────────────────────── */
    .stDataFrame {
        border-radius: 10px !important;
        overflow: hidden !important;
    }

    /* ── Hide Streamlit branding ───────────────────────────────────────────── */
    /* ── User-corrected field badge ──────────────────────────────────────── */
    .user-corrected {
        border-left: 3px solid var(--orange);
        padding-left: 8px;
        background: rgba(251, 155, 53, 0.06);
        border-radius: 0 6px 6px 0;
        margin: 2px 0;
        font-size: 0.73rem;
        color: var(--orange);
        font-weight: 600;
    }

    /* ── Data bar for monthly claims ──────────────────────────────────────── */
    .data-bar-bg {
        background: rgba(53,197,252,0.12);
        border-radius: 4px;
        height: 22px;
        width: 100%;
        overflow: hidden;
    }
    .data-bar-fill {
        height: 100%;
        border-radius: 4px;
        transition: width 0.3s ease;
    }
    .data-bar-fill.blue {
        background: linear-gradient(90deg, #35c5fc, #003780);
    }
    .data-bar-fill.warm {
        background: linear-gradient(90deg, #fb9b35, #f1517b);
    }

    /* ── Monthly claims row header ───────────────────────────────────────── */
    .monthly-header {
        font-family: 'Raleway', sans-serif;
        font-size: 0.68rem;
        font-weight: 800;
        text-transform: uppercase;
        letter-spacing: 1px;
        color: var(--navy);
        padding: 8px 0;
    }

    #MainMenu { visibility: hidden; }
    footer { visibility: hidden; }
    </style>
    """, unsafe_allow_html=True)


def render_sidebar():
    """Render the sidebar navigation styled like the WellX header nav."""
    with st.sidebar:
        # Brand header matching .brand-name + header-text from index.html
        st.markdown("""
        <div style="padding: 8px 0 16px;">
            <div style="font-family:'Raleway',sans-serif; font-weight:900; font-size:0.65rem;
                        letter-spacing:3.5px; text-transform:uppercase;
                        background:linear-gradient(90deg,#fb9b35,#f1517b,#b43082,#8431cb,#35c5fc);
                        -webkit-background-clip:text; -webkit-text-fill-color:transparent;
                        background-clip:text; margin-bottom:6px;">
                WellX
            </div>
            <div style="font-family:'Raleway',sans-serif; font-weight:800; font-size:1.2rem;
                        color:#fff; line-height:1.2; margin-bottom:4px;">
                AI Underwriter
            </div>
            <div style="font-size:0.78rem; color:rgba(255,255,255,0.4); line-height:1.4;">
                DHA Report Analysis &amp; Premium Quotation
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("---")

        nav_options = ["📝 New Quote", "📋 Extracted Information", "🔄 Revisions", "📊 Dashboard"]
        # Auto-select Extracted Information page when extraction is pending
        default_idx = nav_options.index(st.session_state.get("active_page", "📝 New Quote")) \
            if st.session_state.get("active_page") in nav_options else 0

        page = st.radio(
            "Navigation",
            nav_options,
            index=default_idx,
            label_visibility="collapsed",
        )

        st.markdown("---")

        # Underwriter identity — stamped into Excel exports
        st.session_state["prepared_by"] = st.text_input(
            "Prepared by",
            value=st.session_state.get("prepared_by", ""),
            placeholder="e.g. Jasper",
            help="Name shown in the 'Prepared by' footer of the downloaded Excel.",
        )

        st.markdown("---")

        # Footer matching the subtle brand text
        st.markdown("""
        <div style="font-size:0.68rem; color:rgba(255,255,255,0.3); line-height:1.5; padding:8px 0;">
            Powered by <span style="color:rgba(255,255,255,0.5); font-weight:600;">Jasper AI</span>
            &amp; WellX SOP Engine
        </div>
        """, unsafe_allow_html=True)

    return page


def render_metric(label: str, value, sub: str = "", style: str = "", currency: bool = True):
    """Render a styled stat card. Set currency=False for plain numbers."""
    if isinstance(value, (int, float)):
        val_str = f"AED {value:,.2f}" if currency else f"{value:,.0f}" if isinstance(value, int) else f"{value:,.2f}"
    else:
        val_str = str(value)
    css_class = f"stat-card {style}" if style else "stat-card"
    st.markdown(f"""
    <div class="{css_class}">
        <div class="stat-val">{val_str}</div>
        <div class="stat-lbl">{label}</div>
        <div class="stat-sub">{sub}</div>
    </div>
    """, unsafe_allow_html=True)


def display_census_analysis(ca: dict):
    """Display census analysis results (employee/dependent split, age dist, flags)."""

    st.markdown('<div class="section-lbl">Census Analysis</div>', unsafe_allow_html=True)

    # ── Member Type Breakdown ──
    col1, col2, col3 = st.columns(3)
    with col1:
        render_metric(
            "Employees / Principals",
            f"{ca['employees']} ({ca['employee_pct']}%)",
            style="teal",
        )
    with col2:
        render_metric(
            "Dependents",
            f"{ca['dependents']} ({ca['dependent_pct']}%)",
        )
    with col3:
        mf = ca.get("married_females_18_45", 0)
        mf_pct = ca.get("married_females_18_45_pct", 0)
        style = "highlight" if mf_pct > 15 else ""
        render_metric(
            "Married Females (18-45)",
            f"{mf} ({mf_pct}%)",
            sub="Flag if >15%",
            style=style,
        )

    # ── Age Distribution ──
    age_dist = ca.get("age_distribution", {})
    if age_dist:
        st.markdown("**Age Distribution:**")
        age_cols = st.columns(len(age_dist))
        for i, (bracket, info) in enumerate(age_dist.items()):
            with age_cols[i]:
                st.markdown(f"""
                <div class="stat-card" style="padding:10px 8px;">
                    <div class="stat-val" style="font-size:1rem;">{info['count']}</div>
                    <div class="stat-lbl" style="font-size:0.65rem;">{bracket}</div>
                    <div class="stat-sub">{info['pct']}%</div>
                </div>
                """, unsafe_allow_html=True)

    # ── Flags ──
    if ca.get("flags"):
        for flag in ca["flags"]:
            st.markdown(f'<div class="badge-fail">🚩 {flag}</div>', unsafe_allow_html=True)

    if ca.get("disallowed_relations"):
        st.markdown(
            f'<div class="badge-fail">❌ Disallowed relation types found: '
            f'{", ".join(ca["disallowed_relations"])}</div>',
            unsafe_allow_html=True,
        )

    if not ca.get("flags") and not ca.get("disallowed_relations"):
        st.markdown('<div class="badge-ok">✅ Census looks clean — no flags</div>', unsafe_allow_html=True)


def display_summary(summary: dict, data: dict):
    """Display the full analysis summary in a professional WellX-branded layout."""

    # ── Header banner ──
    st.markdown("""
    <div style="background:#000; border-radius:14px; overflow:hidden; margin-bottom:24px;">
        <div style="height:3px; background:linear-gradient(90deg,#fb9b35,#f1517b,#b43082,#8431cb,#35c5fc);"></div>
        <div style="padding:22px 28px;">
            <div style="font-family:'Raleway',sans-serif; font-weight:900; font-size:0.6rem;
                        letter-spacing:3px; text-transform:uppercase;
                        background:linear-gradient(90deg,#fb9b35,#f1517b,#b43082,#8431cb,#35c5fc);
                        -webkit-background-clip:text; -webkit-text-fill-color:transparent;
                        margin-bottom:6px;">WellX AI Underwriter</div>
            <div style="font-family:'Raleway',sans-serif; font-weight:800; font-size:1.3rem;
                        color:#fff;">Analysis Results</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── 1. Validation & Recency ──
    st.markdown('<div class="section-lbl">1 &middot; Validation &amp; Recency</div>', unsafe_allow_html=True)
    v = summary["validations"]

    col1, col2, col3 = st.columns(3)
    with col1:
        match = v.get("employer_match")
        if match:
            st.markdown('<div class="badge-ok">✅ Employer Match Confirmed</div>', unsafe_allow_html=True)
        elif match is False:
            st.markdown('<div class="badge-fail">❌ Employer Mismatch</div>', unsafe_allow_html=True)
    with col2:
        valid = v.get("policy_valid")
        if valid:
            st.markdown(f'<div class="badge-ok">✅ Policy Valid ({v.get("policy_days_since", "?")} days)</div>', unsafe_allow_html=True)
        elif valid is False:
            st.markdown(f'<div class="badge-fail">❌ Policy Expired ({v.get("policy_days_since", "?")} days)</div>', unsafe_allow_html=True)
    with col3:
        recent = v.get("report_recent")
        if recent:
            st.markdown(f'<div class="badge-ok">✅ Report Current ({v.get("report_days_old", "?")} days old)</div>', unsafe_allow_html=True)
        elif recent is False:
            st.markdown(f'<div class="badge-fail">❌ Stale Report ({v.get("report_days_old", "?")} days old)</div>', unsafe_allow_html=True)

    # ── 2. Claims Ratio & Member Type ──
    st.markdown('<div class="section-lbl">2 &middot; Claims Ratio &amp; Member Type</div>', unsafe_allow_html=True)
    ca = summary["claims_analysis"]
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        render_metric("Claims Paid", ca["claims_paid"])
    with col2:
        render_metric("Outstanding", ca["claims_outstanding"], f"Ratio: {ca['outstanding_ratio_pct']}%")
    with col3:
        render_metric("IBNR", ca["claims_ibnr"])
    with col4:
        render_metric("IP Ratio", f"{ca['ip_ratio_pct']}%", f"Benchmark: {ca['ip_benchmark']}", "teal")

    # ── 3. Census & Population Growth ──
    st.markdown('<div class="section-lbl">3 &middot; Census &amp; Population Growth</div>', unsafe_allow_html=True)
    ce = summary["census_analysis"]
    col1, col2, col3 = st.columns(3)
    with col1:
        render_metric("Starting Census", ce["census_start"], currency=False)
    with col2:
        render_metric("Ending Census", ce["census_end"], currency=False)
    with col3:
        change_str = f"{ce['census_change_pct']:+.1f}%"
        flag_style = "highlight" if abs(ce['census_change_pct']) > 15 else ""
        render_metric("Census Change", change_str, f"Benchmark: {ce['benchmark']}", flag_style)

    # ── 4. Diagnosis & Provider Evaluation ──
    st.markdown('<div class="section-lbl">4 &middot; Diagnosis &amp; Provider Evaluation</div>', unsafe_allow_html=True)
    da = summary["diagnosis_analysis"]
    if da["flagged_conditions"]:
        for cond in da["flagged_conditions"]:
            st.markdown(f'<div class="badge-fail">🔴 Major Condition: {cond}</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="badge-ok">✅ No major conditions (neoplasms, cancer, autoimmune) detected</div>', unsafe_allow_html=True)

    if da["high_value_claims"]:
        st.markdown("**High-Value Claims (>30k AED per claim):**")
        hvc_df = pd.DataFrame(da["high_value_claims"])
        st.dataframe(hvc_df, use_container_width=True)

    st.markdown(f"**Top-10 Concentration:** {da['top10_concentration_pct']}% of total claims")

    # ── 5. Burning Cost Calculation ──
    st.markdown('<div class="section-lbl">5 &middot; Burning Cost Calculation</div>', unsafe_allow_html=True)
    bc = summary["burning_cost_analysis"]

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**Three-Average Method:**")
        avg_df = pd.DataFrame({
            "Average": ["A (All incurred)", "B (Excl. last)", "C (Excl. last 2)"],
            "Value (AED)": [f"{bc['avg_a']:,.2f}", f"{bc['avg_b']:,.2f}", f"{bc['avg_c']:,.2f}"],
        })
        st.dataframe(avg_df, use_container_width=True, hide_index=True)
        st.markdown(f"**Method:** {bc['method']}")
        st.markdown(f"**Highest Average (Monthly Burning Cost):** AED {bc['highest_avg_monthly']:,.2f}")

        if bc.get("sum_matches_paid"):
            st.markdown('<div class="badge-ok">✅ Monthly sum matches paid claims</div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="badge-warn">⚠ Monthly sum differs from paid claims by AED {bc["sum_check_vs_paid"]:,.2f}</div>', unsafe_allow_html=True)

    with col2:
        # ── Burning Cost Per Capita formula breakdown ──
        avg_cen = summary["census_analysis"]["avg_census"]
        st.markdown(f"""
        <div class="info-box">
            <strong>Burning Cost Per Capita</strong><br>
            Monthly Burning Cost (<strong>AED {bc['highest_avg_monthly']:,.2f}</strong>)
            / Normalized Census (<strong>{avg_cen:,.2f}</strong>)
            = <strong style="color:#003780;">AED {bc['burning_cost_per_capita']:,.2f}</strong>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("**Adjustments:**")
        adj_data = {
            "Adjustment": ["Inflation", "IP Allowance", "Outstanding Overflow", "Total"],
            "%": [
                f"+{bc['inflation_pct']}%",
                f"+{bc['ip_adjustment_pct']}%",
                f"+{bc['outstanding_adjustment_pct']}%",
                f"+{bc['inflation_pct'] + bc['ip_adjustment_pct'] + bc['outstanding_adjustment_pct']}%",
            ],
        }
        st.dataframe(pd.DataFrame(adj_data), use_container_width=True, hide_index=True)
        render_metric("Burning Cost Per Capita", bc["burning_cost_per_capita"])
        render_metric("Adjusted Burning Cost Per Capita", bc["adjusted_burning_cost_per_capita"])

    # Show monthly claims table
    if data.get("monthly_claims"):
        with st.expander("📅 View Monthly Claims (Section 17)"):
            monthly_df = pd.DataFrame(data["monthly_claims"])
            st.dataframe(monthly_df, use_container_width=True, hide_index=True)

    # ── 6. Premium Quotation ──
    st.markdown('<div class="section-lbl">6 &middot; Premium Quotation</div>', unsafe_allow_html=True)
    pq = summary["premium_quotation"]

    col1, col2, col3 = st.columns(3)
    with col1:
        render_metric("Projected Annual Claims", pq["projected_claims_annual"])
    with col2:
        render_metric(
            "Total Commission",
            f"{pq['total_commission_pct']}%",
            f"Excl. RI Margin: {pq.get('total_excl_ri_margin_pct', 0)}%",
            "teal",
        )
    with col3:
        # Hero premium card
        st.markdown(f"""
        <div class="premium-hero">
            <div class="ph-label">Indicative Premium</div>
            <div class="ph-value">AED {pq['indicative_premium']:,.2f}</div>
            <div class="ph-sub">Per member/year: AED {pq['premium_per_member_annual']:,.2f}</div>
            <div class="ph-sub">Per member/month: AED {pq['premium_per_member_monthly']:,.2f}</div>
        </div>
        """, unsafe_allow_html=True)

    # ── Platform-specific commission calculations ──
    platform = pq.get("platform_key", "HealthX")
    col1, col2, col3 = st.columns(3)
    with col1:
        render_metric(
            f"{platform} Actual Commission",
            pq.get("actual_commission", 0),
            f"{platform} {pq.get('platform_pct', 0)}% of Indicative Premium",
            "highlight",
        )
    with col2:
        min_pct = pq.get("min_platform_pct", 0)
        render_metric(
            f"Minimum {platform} %",
            f"{min_pct:.4f}%",
            f"= (110 x {pq['current_census']}) / Indicative Premium",
            "teal",
        )
    with col3:
        render_metric(
            "Current Census",
            pq["current_census"],
            f"Plan: {pq.get('plan', 'N/A')}",
            currency=False,
        )

    # ── Flags ──
    if summary.get("flags"):
        st.markdown('<div class="section-lbl">Flags &amp; Warnings</div>', unsafe_allow_html=True)
        for flag in summary["flags"]:
            st.markdown(f'<div class="badge-warn">⚠ {flag}</div>', unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# PAGE: NEW QUOTE
# ---------------------------------------------------------------------------

def page_new_quote():
    """Render the New Quote page — input, upload, analyze, and save."""
    # ── Page header matching index.html card-header style ──
    st.markdown("""
    <div style="background:#000; border-radius:14px; overflow:hidden; margin-bottom:28px;">
        <div style="height:3px; background:linear-gradient(90deg,#fb9b35,#f1517b,#b43082,#8431cb,#35c5fc);"></div>
        <div style="padding:22px 28px;">
            <div style="font-family:'Raleway',sans-serif; font-weight:900; font-size:0.6rem;
                        letter-spacing:3px; text-transform:uppercase;
                        background:linear-gradient(90deg,#fb9b35,#f1517b,#b43082,#8431cb,#35c5fc);
                        -webkit-background-clip:text; -webkit-text-fill-color:transparent;
                        margin-bottom:6px;">WellX</div>
            <div style="font-family:'Raleway',sans-serif; font-weight:800; font-size:1.3rem;
                        color:#fff; line-height:1.2; margin-bottom:4px;">
                New Quote
                <span style="display:block; font-size:0.85rem; font-weight:500;
                             color:rgba(255,255,255,0.45); font-family:'Inter',sans-serif; margin-top:4px;">
                    Upload a DHA report or claims data to generate a premium quote
                </span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # API key is hardcoded — no need for user input
    api_key = ANTHROPIC_API_KEY

    # ── File Uploads ──
    st.markdown('<div class="section-lbl">Upload Files</div>', unsafe_allow_html=True)
    col_up1, col_up2 = st.columns(2)
    with col_up1:
        uploaded_file = st.file_uploader(
            "DHA Report (PDF) or Claims Data",
            type=["pdf", "xlsx", "xls", "csv"],
            help="PDF: Analyzed by Jasper AI.  Excel/CSV: Line-by-line claims data.",
            key="dha_upload",
        )
    with col_up2:
        census_files = st.file_uploader(
            "Census File(s) (Excel/CSV)",
            type=["xlsx", "xls", "csv"],
            accept_multiple_files=True,
            help="Upload one or multiple census files. Supports Liva/Nextcare, MaxHealth/Chrome, and standard formats. Files are auto-detected and consolidated.",
            key="census_upload",
        )

    # ── Company, Broker & Plan ──
    st.markdown('<div class="section-lbl">Client &amp; Broker</div>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns(3)
    with col1:
        company_name = st.text_input("Company / Employer Name", placeholder="e.g. United Bank Limited")
    with col2:
        broker_name = st.text_input("Broker Name", placeholder="e.g. Marsh McLennan")
    with col3:
        plan = st.selectbox("Plan", PLAN_OPTIONS, index=0)

    col4, col5 = st.columns(2)
    with col4:
        underwriter = st.selectbox("Underwriter", ["Jasper", "Mabel", "Joseph", "Angela"])
    with col5:
        rm_person = st.selectbox("RM", ["Heston", "Mark", "Sujith"])

    # ── Dynamic Commissions based on selected plan ──
    st.markdown('<div class="section-lbl">Commissions &amp; Margins</div>', unsafe_allow_html=True)
    st.markdown(
        f'<div class="info-box"><strong>Plan: {plan}</strong> — '
        f'Default rates loaded. Adjust as needed.</div>',
        unsafe_allow_html=True,
    )

    defaults = COMMISSION_DEFAULTS[plan]
    comm_keys = list(defaults.keys())

    # Render commission inputs dynamically
    commissions = {}
    cols = st.columns(len(comm_keys))
    for i, key in enumerate(comm_keys):
        with cols[i]:
            val = st.number_input(
                key,
                value=defaults[key],
                min_value=0.0,
                max_value=50.0,
                step=0.5,
                key=f"comm_{plan}_{key}",
            )
            commissions[key] = val

    # Show total and total excluding RI Margin
    total_all = sum(commissions.values())
    ri_margin_val = commissions.get("Reinsurance Margin", 0)
    total_excl_ri = total_all - ri_margin_val
    st.markdown(
        f'<div class="info-box">'
        f'<strong>Total Commission:</strong> {total_all:.1f}% &nbsp;|&nbsp; '
        f'<strong>Total excl. Reinsurance Margin:</strong> {total_excl_ri:.1f}%'
        f'</div>',
        unsafe_allow_html=True,
    )

    st.markdown("---")

    # --- Analyze Button ---
    if st.button("🚀 Analyze & Generate Quote", type="primary", use_container_width=True):
        if not uploaded_file and not census_files:
            st.error("Please upload at least a DHA report or census file.")
            return
        if not company_name:
            st.error("Please enter the Company / Employer name.")
            return

        # ── Census analysis (always run if census file(s) provided) ──
        census_analysis = None
        census_df = None
        if census_files:
            try:
                # Always run through consolidation pipeline — it auto-detects
                # Liva/Nextcare, MaxHealth/Chrome, and standard formats, handles
                # header-row detection, and filters to active members only.
                census_df, consolidation_summary = consolidate_census_files(census_files)
                st.session_state["census_consolidation_summary"] = consolidation_summary

                total_active = consolidation_summary["total_active"]
                total_inactive = consolidation_summary["total_inactive"]
                file_count = len(consolidation_summary["per_file"])
                has_errors = any(fi.get("error") for fi in consolidation_summary["per_file"])

                # Determine if any file used insurer format (vs all standard)
                formats_used = {fi["format"] for fi in consolidation_summary["per_file"] if not fi.get("error")}
                is_insurer_format = bool(formats_used - {"Standard"})

                if is_insurer_format or len(census_files) > 1:
                    st.success(
                        f"Census consolidated: {file_count} file(s) → "
                        f"**{total_active}** active members"
                        + (f" ({total_inactive} inactive excluded)" if total_inactive > 0 else "")
                    )

                    # Show per-file breakdown
                    with st.expander("Census consolidation details", expanded=False):
                        for fi in consolidation_summary["per_file"]:
                            err = fi.get("error")
                            if err:
                                st.error(f"**{fi['name']}**: Failed to read — {err}")
                            else:
                                st.markdown(
                                    f"- **{fi['name']}** — Format: {fi['format']}, "
                                    f"Active: {fi['active']}, Inactive excluded: {fi['inactive']}"
                                )

                # Map columns for analyze_census_file (expects RELATION, GENDER, Date Of Birth)
                analysis_df = census_df.rename(columns={
                    "Relation": "RELATION",
                    "Gender": "GENDER",
                    "DOB": "Date Of Birth",
                })
                census_analysis = analyze_census_file(analysis_df)

                st.session_state["last_census_df"] = census_df
                st.session_state["last_census_analysis"] = census_analysis
            except Exception as e:
                st.error(f"Failed to read census file(s): {e}")

        # ── DHA PDF Analysis ──
        if uploaded_file:
            file_ext = uploaded_file.name.split(".")[-1].lower()

            if file_ext == "pdf":
                pdf_bytes = uploaded_file.read()
                data = extract_dha_report_with_claude(api_key, pdf_bytes)

                if not data:
                    return

                st.success("Report extracted successfully! Redirecting to review...")

                # Store raw extraction and initialize editable copy
                st.session_state["last_extract"] = data
                st.session_state["editable_extract"] = copy.deepcopy(data)
                st.session_state["last_commissions"] = commissions
                st.session_state["last_company"] = company_name
                st.session_state["last_broker"] = broker_name
                st.session_state["last_plan"] = plan
                st.session_state["last_underwriter"] = underwriter
                st.session_state["last_rm"] = rm_person
                st.session_state["user_corrections"] = {}

                # Initialize monthly controls
                monthly = data.get("monthly_claims", [])
                st.session_state["monthly_included"] = [
                    float(m.get("value", 0) or 0) > 0 for m in monthly
                ]
                st.session_state["monthly_haircuts"] = [0.0] * len(monthly)

                # Auto-tick months per SOP: determine which set gives highest average
                # and pre-select those months
                vals = [float(m.get("value", 0) or 0) for m in monthly]
                incurred_idx = [i for i, v in enumerate(vals) if v > 0]
                policy_eff_dt = parse_date_flexible(data.get("policy_effective_date", ""))
                psd = policy_eff_dt.day if policy_eff_dt else 1

                if len(incurred_idx) >= 3:
                    if psd <= 5:
                        sets = [incurred_idx, incurred_idx[:-1], incurred_idx[:-2]]
                    else:
                        sets = [incurred_idx[1:], incurred_idx[1:-1], incurred_idx[1:-2]]
                    avgs = []
                    for s in sets:
                        if s:
                            avgs.append(sum(vals[j] for j in s) / len(s))
                        else:
                            avgs.append(0)
                    best_set = sets[avgs.index(max(avgs))]
                    auto_included = [False] * len(monthly)
                    for j in best_set:
                        auto_included[j] = True
                    st.session_state["monthly_included"] = auto_included

                # Initialize adjustment overrides
                st.session_state["adj_inflation"] = 5.0
                st.session_state["adj_ip"] = None  # None = auto-calculate
                st.session_state["adj_os"] = None
                st.session_state["major_claims_allowance"] = 0.0
                st.session_state["uw_loading_pct"] = 0.0
                st.session_state["uw_discount_pct"] = 0.0
                st.session_state["uw_loading_note"] = ""
                st.session_state["uw_discount_note"] = ""

                # Redirect to Extracted Information page
                st.session_state["active_page"] = "📋 Extracted Information"
                st.rerun()

            elif file_ext in ("xlsx", "xls", "csv"):
                try:
                    if file_ext == "csv":
                        df = pd.read_csv(uploaded_file)
                    else:
                        df = pd.read_excel(uploaded_file)
                except Exception as e:
                    st.error(f"Failed to read file: {e}")
                    return

                st.success(f"Loaded {len(df)} rows from {uploaded_file.name}")
                st.dataframe(df.head(20), use_container_width=True)
                st.session_state["last_census_df"] = df
                return

        elif census_files and census_analysis:
            # Census-only mode (no DHA report)
            display_census_analysis(census_analysis)
            return

    # --- Post-Analysis Actions (only shown after analysis completes) ---
    if "last_summary" in st.session_state and "last_extract" in st.session_state:
        st.markdown("---")
        st.markdown('<div class="section-lbl">Actions</div>', unsafe_allow_html=True)

        col1, col2, col3 = st.columns(3)

        # Download Excel
        with col1:
            summary = st.session_state["last_summary"]
            data = st.session_state["last_extract"]
            comms = st.session_state.get("last_commissions", commissions)

            excel_bytes = generate_quote_excel(
                summary,
                data,
                comms,
                prepared_by=st.session_state.get("prepared_by", ""),
                broker_name=st.session_state.get("last_broker", ""),
                rm_name=st.session_state.get("last_rm", ""),
            )

            st.download_button(
                label="📥 Download Full Quote Excel",
                data=excel_bytes,
                file_name=f"Claims_Analysis_{summary.get('company_name', 'quote')}_{datetime.now().strftime('%Y%b%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        # Save to database
        with col2:
            status = st.selectbox(
                "Quote Status",
                ["neutral", "positive", "not good", "will confirm", "confirmed", "lost"],
                index=0,
            )

        with col3:
            if st.button("💾 Save Quote to Database", use_container_width=True):
                # Map dynamic commission keys to DB columns
                comm_broker = comms.get("Broker", comms.get("broker", 10))
                comm_insurer = comms.get("Insurance Tax", comms.get("insurer", 0.5))
                comm_tpa = comms.get("NAS", comms.get("tpa", 4))
                comm_wellx = comms.get("HealthX", comms.get("OpenX", comms.get("wellx", 4)))
                comm_margins = comms.get("Reinsurance Margin", comms.get("margins", 7))
                quote_data = {
                    "company_name": st.session_state.get("last_company", ""),
                    "broker_name": st.session_state.get("last_broker", ""),
                    "status": status,
                    "summary": summary,
                    "raw_extract": data,
                    "commission_broker": comm_broker,
                    "commission_insurer": comm_insurer,
                    "commission_tpa": comm_tpa,
                    "commission_wellx": comm_wellx,
                    "commission_margins": comm_margins,
                    "burning_cost": summary["burning_cost_analysis"]["adjusted_burning_cost_per_capita"],
                    "indicative_premium": summary["premium_quotation"]["indicative_premium"],
                    "current_census": summary["premium_quotation"]["current_census"],
                }
                quote_id = save_quote(quote_data)
                st.success(f"Quote saved! (ID: {quote_id})")

        # View raw extraction
        with st.expander("🔍 View Raw Extracted Data (JSON)"):
            st.json(st.session_state["last_extract"])


# ---------------------------------------------------------------------------
# PAGE: EXTRACTED INFORMATION — review, edit, haircut, and confirm
# ---------------------------------------------------------------------------

def calculate_live_premium(data: dict, commissions: dict, included: list, haircuts: list,
                           census_count: int, inflation_pct: float = 5.0,
                           ip_adj_pct: float = None, os_adj_pct: float = None,
                           major_claims_allowance: float = 0.0,
                           uw_loading_pct: float = 0.0, uw_discount_pct: float = 0.0) -> dict:
    """
    Live premium calculator with editable adjustments, major claims allowance,
    and UW loading/discount. Runs on every Streamlit rerun for instant feedback.
    """
    monthly = data.get("monthly_claims", [])

    # Build net values for included months only
    net_values = []
    for i, m in enumerate(monthly):
        val = float(m.get("value", 0) or 0)
        if i < len(included) and included[i] and val > 0:
            net = max(val - (haircuts[i] if i < len(haircuts) else 0), 0)
            net_values.append(net)

    n_incurred = len(net_values)

    # Determine policy start day
    policy_eff = parse_date_flexible(data.get("policy_effective_date", ""))
    policy_start_day = policy_eff.day if policy_eff else 1

    # Three-average method
    if n_incurred >= 3:
        if policy_start_day <= 5:
            avg_a = sum(net_values) / len(net_values)
            avg_b = sum(net_values[:-1]) / len(net_values[:-1]) if len(net_values) > 1 else avg_a
            avg_c = sum(net_values[:-2]) / len(net_values[:-2]) if len(net_values) > 2 else avg_b
        else:
            excl = net_values[1:]
            avg_a = sum(excl) / len(excl) if excl else 0
            excl2 = net_values[1:-1]
            avg_b = sum(excl2) / len(excl2) if excl2 else avg_a
            excl3 = net_values[1:-2]
            avg_c = sum(excl3) / len(excl3) if excl3 else avg_b
    elif n_incurred > 0:
        avg_a = avg_b = avg_c = sum(net_values) / len(net_values)
    else:
        avg_a = avg_b = avg_c = 0

    highest_avg = max(avg_a, avg_b, avg_c)

    # Census from DHA report for burning cost denominator
    def sum_census_inner(census_data):
        if not census_data:
            return 0
        total = census_data.get("grand_total", 0)
        if total:
            return int(total)
        s = 0
        for cat in ("male", "single_female", "married_female"):
            cat_data = census_data.get(cat, {})
            if isinstance(cat_data, dict):
                cat_total = cat_data.get("total", 0)
                if cat_total:
                    s += int(cat_total)
                else:
                    for k, v in cat_data.items():
                        if k != "total":
                            s += int(v or 0)
            elif isinstance(cat_data, (int, float)):
                s += int(cat_data)
        return s

    cs = sum_census_inner(data.get("census_start"))
    ce = sum_census_inner(data.get("census_end"))
    avg_census = (cs + ce) / 2 if (cs + ce) > 0 else 1

    burning_cost = highest_avg / avg_census if avg_census > 0 else 0

    # Adjustments — use overrides if provided
    claims_paid = float(data.get("claims_paid", 0) or 0)
    claims_outstanding = float(data.get("claims_outstanding", 0) or 0)
    member_type = data.get("claims_by_member_type", {})
    totals_row = member_type.get("totals", member_type.get("Totals", {}))
    ip_total = float(totals_row.get("ip", 0) or 0)
    claims_total_s8 = float(totals_row.get("total", 0) or 0)
    ip_ratio = (ip_total / claims_total_s8 * 100) if claims_total_s8 > 0 else 0
    outstanding_ratio = (claims_outstanding / claims_paid * 100) if claims_paid > 0 else 0

    inflation = inflation_pct / 100
    if ip_adj_pct is not None:
        ip_adj = ip_adj_pct / 100
    else:
        ip_adj = (25 - ip_ratio) / 100 if ip_ratio < 20 else 0
    if os_adj_pct is not None:
        out_adj = os_adj_pct / 100
    else:
        out_adj = (outstanding_ratio - 20) / 100 if outstanding_ratio > 20 else 0

    adjusted = burning_cost * (1 + inflation + ip_adj + out_adj)

    # Premium calculation
    current = census_count if census_count > 0 else (ce if ce > 0 else cs)
    projected = adjusted * 12 * current  # Net premium before loading/discount
    total_comm = sum(commissions.values()) / 100

    # Apply UW loading or discount on net premium first
    net_after_uw = projected * (1 + uw_loading_pct / 100 - uw_discount_pct / 100)
    # Then add major claims allowance
    adjusted_net = net_after_uw + major_claims_allowance

    indicative = adjusted_net / (1 - total_comm) if total_comm < 1 else adjusted_net

    return {
        "avg_a": round(avg_a, 2),
        "avg_b": round(avg_b, 2),
        "avg_c": round(avg_c, 2),
        "highest_avg": round(highest_avg, 2),
        "avg_census": round(avg_census, 2),
        "census_start": cs,
        "census_end": ce,
        "burning_cost": round(burning_cost, 2),
        "adjusted_bc": round(adjusted, 2),
        "projected": round(projected, 2),
        "net_after_uw": round(net_after_uw, 2),
        "major_claims_allowance": round(major_claims_allowance, 2),
        "adjusted_net": round(adjusted_net, 2),
        "indicative": round(indicative, 2),
        "current_census": current,
        "n_months": n_incurred,
        "ip_ratio": round(ip_ratio, 2),
        "outstanding_ratio": round(outstanding_ratio, 2),
        "inflation_pct": inflation_pct,
        "ip_adj_pct": round(ip_adj * 100, 2),
        "os_adj_pct": round(out_adj * 100, 2),
    }


def page_extracted_info():
    """Review extracted data, edit corrections, set haircuts, and confirm analysis."""

    # ── Guard ──
    if "editable_extract" not in st.session_state:
        st.markdown("""
        <div class="info-box">
            <strong>No extraction data available.</strong><br>
            Go to <strong>New Quote</strong> and upload a DHA report to get started.
        </div>
        """, unsafe_allow_html=True)
        return

    data = st.session_state["editable_extract"]
    original = st.session_state.get("last_extract", {})
    corrections = st.session_state.get("user_corrections", {})

    # ── Page header ──
    st.markdown("""
    <div style="background:#000; border-radius:14px; overflow:hidden; margin-bottom:28px;">
        <div style="height:3px; background:linear-gradient(90deg,#fb9b35,#f1517b,#b43082,#8431cb,#35c5fc);"></div>
        <div style="padding:22px 28px;">
            <div style="font-family:'Raleway',sans-serif; font-weight:900; font-size:0.6rem;
                        letter-spacing:3px; text-transform:uppercase;
                        background:linear-gradient(90deg,#fb9b35,#f1517b,#b43082,#8431cb,#35c5fc);
                        -webkit-background-clip:text; -webkit-text-fill-color:transparent;
                        margin-bottom:6px;">WellX</div>
            <div style="font-family:'Raleway',sans-serif; font-weight:800; font-size:1.3rem;
                        color:#fff; line-height:1.2; margin-bottom:4px;">
                Extracted Information
                <span style="display:block; font-size:0.85rem; font-weight:500;
                             color:rgba(255,255,255,0.45); font-family:'Inter',sans-serif; margin-top:4px;">
                    Review AI extraction, correct errors, and set claims haircuts
                </span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Helper to track corrections ──
    def track_correction(field_key, new_val):
        orig_val = original.get(field_key, "")
        if str(new_val) != str(orig_val):
            corrections[field_key] = {"original": orig_val, "corrected": new_val}
        elif field_key in corrections:
            del corrections[field_key]
        st.session_state["user_corrections"] = corrections

    # =======================================================================
    # SECTION A: Employer & Dates
    # =======================================================================
    st.markdown('<div class="section-lbl">Employer &amp; Policy Dates</div>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        emp = st.text_input("Employer Name", value=data.get("employer_name", ""), key="ei_employer")
        data["employer_name"] = emp
        track_correction("employer_name", emp)
        if "employer_name" in corrections:
            st.markdown(f'<div class="user-corrected">✏️ Corrected (was: {corrections["employer_name"]["original"]})</div>', unsafe_allow_html=True)

    with col2:
        notes = data.get("extraction_notes", "")
        if notes:
            st.markdown(f'<div class="info-box">{notes}</div>', unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3)
    date_fields = [
        ("policy_effective_date", "Policy Effective Date"),
        ("policy_expiry_date", "Policy Expiry Date"),
        ("report_period_end", "Report Period End"),
    ]
    for (key, label), col in zip(date_fields, [col1, col2, col3]):
        with col:
            val = st.text_input(label, value=data.get(key, ""), key=f"ei_{key}")
            data[key] = val
            track_correction(key, val)
            if key in corrections:
                st.markdown(f'<div class="user-corrected">✏️ Corrected</div>', unsafe_allow_html=True)

    # =======================================================================
    # SECTION B: Claims Values
    # =======================================================================
    st.markdown('<div class="section-lbl">Claims Values (AED)</div>', unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3)
    claims_fields = [
        ("claims_paid", "Claims Paid"),
        ("claims_outstanding", "Claims Outstanding"),
        ("claims_ibnr", "Claims IBNR"),
    ]
    for (key, label), col in zip(claims_fields, [col1, col2, col3]):
        with col:
            val = st.number_input(
                label,
                value=float(data.get(key, 0) or 0),
                min_value=0.0,
                step=1000.0,
                key=f"ei_{key}",
            )
            data[key] = val
            track_correction(key, val)
            if key in corrections:
                try:
                    orig_display = f'{float(corrections[key]["original"]):,.0f}'
                except (ValueError, TypeError):
                    orig_display = corrections[key]["original"]
                st.markdown(f'<div class="user-corrected">✏️ Corrected (was: {orig_display})</div>', unsafe_allow_html=True)

    # =======================================================================
    # SECTION C: Monthly Claims with Checkboxes, Haircuts, Data Bars
    # =======================================================================
    st.markdown('<div class="section-lbl">Monthly Claims (Section 17)</div>', unsafe_allow_html=True)

    monthly = data.get("monthly_claims", [])
    included = st.session_state.get("monthly_included", [True] * len(monthly))
    haircuts = st.session_state.get("monthly_haircuts", [0.0] * len(monthly))

    if monthly:
        # Column headers
        hdr = st.columns([0.6, 1.5, 2.0, 2.5, 2.0, 2.0, 2.5])
        headers = ["Include", "Month", "Claims (AED)", "Claims Bar", "Haircut (AED)", "Net Claims", "Net Bar"]
        for h, col in zip(headers, hdr):
            with col:
                st.markdown(f'<div class="monthly-header">{h}</div>', unsafe_allow_html=True)

        # Compute max for data bar scaling
        all_vals = [float(m.get("value", 0) or 0) for m in monthly]
        max_val = max(all_vals) if all_vals else 1

        for i, m in enumerate(monthly):
            cols = st.columns([0.6, 1.5, 2.0, 2.5, 2.0, 2.0, 2.5])
            val = float(m.get("value", 0) or 0)

            with cols[0]:
                inc = st.checkbox(
                    "Inc", value=included[i] if i < len(included) else True,
                    key=f"mi_{i}", label_visibility="collapsed",
                )
                st.session_state["monthly_included"][i] = inc

            with cols[1]:
                st.markdown(f"**{m.get('month', '')}** {m.get('year', '')}")

            with cols[2]:
                new_val = st.number_input(
                    "val", value=val, min_value=0.0, step=1000.0,
                    key=f"mv_{i}", label_visibility="collapsed",
                )
                data["monthly_claims"][i]["value"] = new_val
                # Track correction
                orig_m = original.get("monthly_claims", [])
                if i < len(orig_m) and new_val != float(orig_m[i].get("value", 0) or 0):
                    corrections[f"monthly_{i}"] = {
                        "original": orig_m[i].get("value", 0),
                        "corrected": new_val,
                    }
                val = new_val  # Use updated value

            with cols[3]:
                pct = (val / max_val * 100) if max_val > 0 else 0
                st.markdown(
                    f'<div class="data-bar-bg">'
                    f'<div class="data-bar-fill blue" style="width:{pct:.1f}%;"></div></div>',
                    unsafe_allow_html=True,
                )

            with cols[4]:
                hc = st.number_input(
                    "hc", value=haircuts[i] if i < len(haircuts) else 0.0,
                    min_value=0.0, step=1000.0,
                    key=f"mh_{i}", label_visibility="collapsed",
                )
                st.session_state["monthly_haircuts"][i] = hc

            with cols[5]:
                net = max(val - hc, 0)
                color = "#155724" if inc else "#9aa5b4"
                strike = "" if inc else "text-decoration:line-through;"
                st.markdown(
                    f'<div style="font-weight:700; color:{color}; {strike} padding-top:8px;">'
                    f'AED {net:,.0f}</div>',
                    unsafe_allow_html=True,
                )

            with cols[6]:
                net_pct = (net / max_val * 100) if max_val > 0 else 0
                bar_class = "warm" if inc else "blue"
                opacity = "1" if inc else "0.3"
                st.markdown(
                    f'<div class="data-bar-bg" style="opacity:{opacity};">'
                    f'<div class="data-bar-fill {bar_class}" style="width:{max(net_pct, 0):.1f}%;"></div></div>',
                    unsafe_allow_html=True,
                )

    # =======================================================================
    # SECTION D: Census Summary (Report + Uploaded)
    # =======================================================================
    st.markdown('<div class="section-lbl">Census Summary</div>', unsafe_allow_html=True)

    census_analysis = st.session_state.get("last_census_analysis")
    census_count = census_analysis["total_members"] if census_analysis else 0

    # Compute report census from extracted data
    def _sum_census(census_data):
        if not census_data:
            return 0
        total = census_data.get("grand_total", 0)
        if total:
            return int(total)
        s = 0
        for cat in ("male", "single_female", "married_female"):
            cat_data = census_data.get(cat, {})
            if isinstance(cat_data, dict):
                ct = cat_data.get("total", 0)
                s += int(ct) if ct else sum(int(v or 0) for k, v in cat_data.items() if k != "total")
            elif isinstance(cat_data, (int, float)):
                s += int(cat_data)
        return s

    report_start = _sum_census(data.get("census_start"))
    report_end = _sum_census(data.get("census_end"))

    # --- Census fallback for non-standard reports ---
    census_assumed = False
    if report_start == 0 and report_end == 0:
        # Try MaxHealth-style: total_members + membership_change_pct
        total_members = data.get("total_members", 0)
        change_pct = data.get("membership_change_pct", None)
        if total_members and total_members > 0:
            if change_pct is not None:
                # total_members is the starting census; apply % change to get ending
                report_start = int(total_members)
                report_end = int(round(total_members * (1 + change_pct / 100)))
            else:
                # Only total_members, no change info — use as ending census
                report_end = int(total_members)
        # If still no census, assume starting = uploaded census * 0.95
        if report_start == 0 and report_end == 0 and census_count > 0:
            report_start = int(round(census_count * 0.95))
            census_assumed = True
        # If we have starting but no ending, prorate using uploaded census + days elapsed
        if report_start > 0 and report_end == 0 and census_count > 0:
            start_str = data.get("policy_effective_date") or data.get("report_period_start", "")
            end_str = data.get("report_period_end", "")
            expiry_str = data.get("policy_expiry_date", "")
            rp_start = parse_date_flexible(start_str)
            rp_end = parse_date_flexible(end_str)
            pol_expiry = parse_date_flexible(expiry_str)
            if rp_start and rp_end and pol_expiry and pol_expiry > rp_start:
                total_days = (pol_expiry - rp_start).days
                elapsed_days = (rp_end - rp_start).days
                if total_days > 0:
                    report_end = int(round(
                        report_start + (census_count - report_start) * elapsed_days / total_days
                    ))
                else:
                    report_end = census_count
            else:
                report_end = census_count
        # Persist derived census back to data so calculations pick it up
        if report_start > 0:
            data["census_start"] = {"grand_total": report_start}
        if report_end > 0:
            data["census_end"] = {"grand_total": report_end}

    col1, col2, col3 = st.columns(3)
    with col1:
        lbl_start = "Report Starting Census"
        if census_assumed:
            lbl_start += " (Assumed)"
        render_metric(lbl_start, report_start, currency=False)
    with col2:
        render_metric("Report Ending Census", report_end, currency=False)
    with col3:
        render_metric("Uploaded Census", census_count if census_count > 0 else "N/A", currency=False)
    if census_assumed:
        st.markdown(
            '<div class="info-box">⚠️ No census data found in report. '
            'Starting census assumed as 5% lower than uploaded census. '
            'Ending census prorated by elapsed policy days.</div>',
            unsafe_allow_html=True,
        )

    # Uploaded census age distribution
    if census_analysis and census_analysis.get("age_distribution"):
        st.markdown("**Uploaded Census Breakdown:**")
        ad = census_analysis["age_distribution"]
        # Remap to required brackets: 0-18, 19-30, 31-40, 41-50, 51-65, Married F (18-45)
        bracket_data = {}
        for bracket, info in ad.items():
            bracket_data[bracket] = info

        # Show employee/dependent + married females + age brackets
        items = []
        items.append(f"**Employees:** {census_analysis.get('employees', 0)} ({census_analysis.get('employee_pct', 0)}%)")
        items.append(f"**Dependents:** {census_analysis.get('dependents', 0)} ({census_analysis.get('dependent_pct', 0)}%)")
        mf = census_analysis.get("married_females_18_45", 0)
        mf_pct = census_analysis.get("married_females_18_45_pct", 0)
        items.append(f"**Married Females (18-45):** {mf} ({mf_pct}%)")

        col_items = st.columns(3)
        for i, item in enumerate(items):
            with col_items[i]:
                st.markdown(item)

        # Age distribution cards
        if bracket_data:
            age_cols = st.columns(len(bracket_data))
            for idx, (bracket, info) in enumerate(bracket_data.items()):
                with age_cols[idx]:
                    st.markdown(f"""
                    <div class="stat-card" style="padding:8px 6px;">
                        <div class="stat-val" style="font-size:0.95rem;">{info['count']}</div>
                        <div class="stat-lbl" style="font-size:0.6rem;">{bracket}</div>
                        <div class="stat-sub">{info['pct']}%</div>
                    </div>
                    """, unsafe_allow_html=True)

    # =======================================================================
    # SECTION D2: Complex Cases (Section 19)
    # =======================================================================
    complex_notes = data.get("complex_cases_notes", "")
    if complex_notes:
        st.markdown('<div class="section-lbl">Section 19 &mdash; Complex Cases</div>', unsafe_allow_html=True)
        st.markdown(
            f'<div class="badge-fail">🔴 <strong>COMPLEX CASES DETECTED — Flag to Underwriter</strong></div>',
            unsafe_allow_html=True,
        )
        st.markdown(f'<div class="info-box">{complex_notes}</div>', unsafe_allow_html=True)

    # =======================================================================
    # SECTION D3: Diagnosis Top 10
    # =======================================================================
    diag_vals = data.get("diagnosis_top10_values", [])
    if diag_vals:
        with st.expander("📊 Diagnosis Top 10 (editable)"):
            diag_df = pd.DataFrame(diag_vals)
            edited_diag = st.data_editor(diag_df, use_container_width=True, key="ei_diag")
            data["diagnosis_top10_values"] = edited_diag.to_dict("records")

    # =======================================================================
    # SECTION E: Adjustments (editable)
    # =======================================================================
    st.markdown('<div class="section-lbl">Adjustments</div>', unsafe_allow_html=True)

    commissions = st.session_state.get("last_commissions", {})

    # First pass to get auto-calculated values for display
    claims_paid_val = float(data.get("claims_paid", 0) or 0)
    claims_outstanding_val = float(data.get("claims_outstanding", 0) or 0)
    member_type_data = data.get("claims_by_member_type", {})
    totals_r = member_type_data.get("totals", member_type_data.get("Totals", {}))
    ip_t = float(totals_r.get("ip", 0) or 0)
    claims_t = float(totals_r.get("total", 0) or 0)
    auto_ip_ratio = (ip_t / claims_t * 100) if claims_t > 0 else 0
    auto_ip_adj = (25 - auto_ip_ratio) if auto_ip_ratio < 20 else 0
    auto_os_ratio = (claims_outstanding_val / claims_paid_val * 100) if claims_paid_val > 0 else 0
    auto_os_adj = (auto_os_ratio - 20) if auto_os_ratio > 20 else 0

    col1, col2, col3 = st.columns(3)
    with col1:
        adj_inflation = st.number_input(
            "Inflation %", value=st.session_state.get("adj_inflation", 5.0),
            min_value=0.0, max_value=50.0, step=0.5, key="ei_inflation",
        )
        st.session_state["adj_inflation"] = adj_inflation
    with col2:
        adj_ip = st.number_input(
            f"IP Allowance % (auto: {auto_ip_adj:.2f}%)",
            value=float(st.session_state.get("adj_ip") if st.session_state.get("adj_ip") is not None else auto_ip_adj),
            min_value=0.0, max_value=50.0, step=0.5, key="ei_ip_adj",
        )
        st.session_state["adj_ip"] = adj_ip
    with col3:
        adj_os = st.number_input(
            f"OS Overflow % (auto: {auto_os_adj:.2f}%)",
            value=float(st.session_state.get("adj_os") if st.session_state.get("adj_os") is not None else auto_os_adj),
            min_value=0.0, max_value=100.0, step=0.5, key="ei_os_adj",
        )
        st.session_state["adj_os"] = adj_os

    # =======================================================================
    # SECTION E2: Major Claims Allowance & UW Loading/Discount
    # =======================================================================
    st.markdown('<div class="section-lbl">Major Claims Allowance &amp; UW Loading/Discount</div>', unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3)
    with col1:
        major_claims = st.number_input(
            "Major Claims Allowance (AED)",
            value=st.session_state.get("major_claims_allowance", 0.0),
            min_value=0.0, step=1000.0, key="ei_major_claims",
            help="Manual entry for unaccounted claims (newly diagnosed, newly added high-cost members)",
        )
        st.session_state["major_claims_allowance"] = major_claims

    with col2:
        uw_loading = st.number_input(
            "UW Loading %",
            value=st.session_state.get("uw_loading_pct", 0.0),
            min_value=0.0, max_value=100.0, step=0.5, format="%.2f", key="ei_uw_loading",
        )
        st.session_state["uw_loading_pct"] = uw_loading
        loading_note = st.text_input(
            "Loading reason", value=st.session_state.get("uw_loading_note", ""),
            placeholder="e.g. High-risk industry", key="ei_loading_note",
        )
        st.session_state["uw_loading_note"] = loading_note

    with col3:
        uw_discount = st.number_input(
            "UW Discount %",
            value=st.session_state.get("uw_discount_pct", 0.0),
            min_value=0.0, max_value=100.0, step=0.5, format="%.2f", key="ei_uw_discount",
        )
        st.session_state["uw_discount_pct"] = uw_discount
        discount_note = st.text_input(
            "Discount reason", value=st.session_state.get("uw_discount_note", ""),
            placeholder="e.g. Reinsurer approved haircut", key="ei_discount_note",
        )
        st.session_state["uw_discount_note"] = discount_note

    # Formula explanation
    st.markdown(
        '<div class="info-box">'
        '<strong>Formula:</strong> (Net Premium × (1 + Loading% − Discount%)) + Major Claims Allowance = Adjusted Net Premium<br>'
        'Indicative Premium = Adjusted Net Premium / (1 − Total Commission%)'
        '</div>',
        unsafe_allow_html=True,
    )

    # =======================================================================
    # SECTION F: Live Premium Calculator
    # =======================================================================
    st.markdown('<div class="section-lbl">Live Premium Estimate</div>', unsafe_allow_html=True)

    live = calculate_live_premium(
        data, commissions,
        st.session_state.get("monthly_included", []),
        st.session_state.get("monthly_haircuts", []),
        census_count,
        inflation_pct=adj_inflation,
        ip_adj_pct=adj_ip,
        os_adj_pct=adj_os,
        major_claims_allowance=major_claims,
        uw_loading_pct=uw_loading,
        uw_discount_pct=uw_discount,
    )

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        render_metric("Months Used", live["n_months"], currency=False)
    with col2:
        render_metric("Monthly Burning Cost", live["highest_avg"])
    with col3:
        render_metric("Burning Cost/Capita", live["burning_cost"])
    with col4:
        render_metric("Adjusted BC/Capita", live["adjusted_bc"])

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        render_metric("Net Premium", live["projected"])
    with col2:
        render_metric("After UW Adj", live["net_after_uw"],
                      f"+{uw_loading:.2f}% / -{uw_discount:.2f}%")
    with col3:
        render_metric("+ Major Claims", live["major_claims_allowance"])
    with col4:
        st.markdown(f"""
        <div class="premium-hero" style="padding:16px 18px;">
            <div class="ph-label" style="font-size:0.65rem;">Live Indicative Premium</div>
            <div class="ph-value" style="font-size:1.3rem;">AED {live['indicative']:,.2f}</div>
            <div class="ph-sub">Census: {live['current_census']} | Avg: {live['avg_census']}</div>
        </div>
        """, unsafe_allow_html=True)

    # =======================================================================
    # SECTION G: Correction Summary
    # =======================================================================
    if corrections:
        st.markdown(
            f'<div class="badge-warn">✏️ {len(corrections)} field(s) manually corrected by underwriter</div>',
            unsafe_allow_html=True,
        )

    # =======================================================================
    # SECTION G: Confirm & Analyze
    # =======================================================================
    st.markdown("---")

    if st.button("✅ Confirm & Run Full Analysis", type="primary", use_container_width=True):
        # Build modified monthly claims with haircuts and exclusions applied
        final_monthly = []
        for i, m in enumerate(data.get("monthly_claims", [])):
            val = float(m.get("value", 0) or 0)
            inc = st.session_state["monthly_included"][i] if i < len(st.session_state.get("monthly_included", [])) else True
            hc = st.session_state["monthly_haircuts"][i] if i < len(st.session_state.get("monthly_haircuts", [])) else 0
            if inc:
                net = max(val - hc, 0)
            else:
                net = 0  # Excluded months → 0 so they're filtered out
            final_monthly.append({"month": m.get("month"), "year": m.get("year"), "value": net})

        analysis_data = copy.deepcopy(data)
        analysis_data["monthly_claims"] = final_monthly

        company = st.session_state.get("last_company", "")
        plan = st.session_state.get("last_plan", "HealthX-QIC")
        uploaded_count = census_count

        summary = run_sop_analysis(analysis_data, commissions, company, plan, uploaded_count)
        st.session_state["last_summary"] = summary
        st.session_state["active_page"] = "📋 Extracted Information"

        # Display census analysis if available
        if census_analysis:
            display_census_analysis(census_analysis)

        # Display full results
        display_summary(summary, analysis_data)

        # ── Post-analysis actions ──
        st.markdown('<div class="section-lbl">Actions</div>', unsafe_allow_html=True)
        col1, col2, col3 = st.columns(3)

        with col1:
            excel_bytes = generate_quote_excel(
                summary,
                analysis_data,
                commissions,
                prepared_by=st.session_state.get("prepared_by", ""),
                broker_name=st.session_state.get("last_broker", ""),
                rm_name=st.session_state.get("last_rm", ""),
            )
            st.download_button(
                label="📥 Download Full Quote Excel",
                data=excel_bytes,
                file_name=f"Claims_Analysis_{company}_{datetime.now().strftime('%Y%b%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        with col2:
            status = st.selectbox(
                "Quote Status",
                ["neutral", "positive", "not good", "will confirm", "confirmed", "lost"],
                index=0,
                key="ei_status",
            )

        with col3:
            if st.button("💾 Save Quote", use_container_width=True, key="ei_save"):
                comm_broker = commissions.get("Broker", 10)
                comm_insurer = commissions.get("Insurance Tax", 0.5)
                comm_tpa = commissions.get("NAS", 4)
                comm_wellx = commissions.get("HealthX", commissions.get("OpenX", 4))
                comm_margins = commissions.get("Reinsurance Margin", 7)
                quote_data = {
                    "company_name": company,
                    "broker_name": st.session_state.get("last_broker", ""),
                    "status": status,
                    "summary": summary,
                    "raw_extract": analysis_data,
                    "commission_broker": comm_broker,
                    "commission_insurer": comm_insurer,
                    "commission_tpa": comm_tpa,
                    "commission_wellx": comm_wellx,
                    "commission_margins": comm_margins,
                    "burning_cost": summary["burning_cost_analysis"]["adjusted_burning_cost_per_capita"],
                    "indicative_premium": summary["premium_quotation"]["indicative_premium"],
                    "current_census": summary["premium_quotation"]["current_census"],
                }
                quote_id = save_quote(quote_data)
                st.success(f"Quote saved! (ID: {quote_id})")


# ---------------------------------------------------------------------------
# PAGE: REVISIONS (v2)
# ---------------------------------------------------------------------------

def page_revisions():
    """Load a previous quote, adjust inputs, and re-calculate."""
    st.markdown("""
    <div style="background:#000; border-radius:14px; overflow:hidden; margin-bottom:28px;">
        <div style="height:3px; background:linear-gradient(90deg,#fb9b35,#f1517b,#b43082,#8431cb,#35c5fc);"></div>
        <div style="padding:22px 28px;">
            <div style="font-family:'Raleway',sans-serif; font-weight:900; font-size:0.6rem;
                        letter-spacing:3px; text-transform:uppercase;
                        background:linear-gradient(90deg,#fb9b35,#f1517b,#b43082,#8431cb,#35c5fc);
                        -webkit-background-clip:text; -webkit-text-fill-color:transparent;
                        margin-bottom:6px;">WellX</div>
            <div style="font-family:'Raleway',sans-serif; font-weight:800; font-size:1.3rem;
                        color:#fff; line-height:1.2; margin-bottom:4px;">
                Revisions
                <span style="display:block; font-size:0.85rem; font-weight:500;
                             color:rgba(255,255,255,0.45); font-family:'Inter',sans-serif; margin-top:4px;">
                    Load a previous quote and adjust census, commissions, or benefits
                </span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    quotes = get_all_quotes()
    if not quotes:
        st.info("No saved quotes yet. Create a new quote first.")
        return

    # Quote selector
    quote_options = {
        f"#{q['id']} — {q['company_name']} ({q['created_at'][:10]})": q["id"]
        for q in quotes
    }
    selected_label = st.selectbox("Select a previous quote to revise", list(quote_options.keys()))
    selected_id = quote_options[selected_label]

    quote = get_quote_by_id(selected_id)
    if not quote:
        st.error("Quote not found.")
        return

    # Load stored data
    try:
        stored_summary = json.loads(quote["summary_json"]) if quote["summary_json"] else {}
        stored_extract = json.loads(quote["raw_extract"]) if quote["raw_extract"] else {}
    except json.JSONDecodeError:
        stored_summary = {}
        stored_extract = {}

    st.markdown(f"**Company:** {quote['company_name']}  |  **Broker:** {quote['broker_name']}  |  **Status:** {quote['status']}")
    st.markdown("---")

    # Revision inputs
    st.markdown('<div class="section-lbl">Adjust Parameters</div>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        new_census = st.number_input(
            "Updated Current Census",
            value=int(quote.get("current_census", 0) or 0),
            min_value=1,
            step=1,
        )
    with col2:
        manual_adjustment_pct = st.number_input(
            "Manual Loading / Discount (%)",
            value=0.0,
            min_value=-50.0,
            max_value=100.0,
            step=1.0,
            help="Positive = loading, Negative = discount. Applied to net premium.",
        )

    st.markdown('<div class="section-lbl">Commissions</div>', unsafe_allow_html=True)
    c1, c2, c3, c4, c5 = st.columns(5)
    with c1:
        rev_broker = st.number_input("Broker ", value=float(quote.get("commission_broker", 10)), step=0.5, key="rev_b")
    with c2:
        rev_insurer = st.number_input("Insurer ", value=float(quote.get("commission_insurer", 0.5)), step=0.1, key="rev_i")
    with c3:
        rev_tpa = st.number_input("TPA ", value=float(quote.get("commission_tpa", 4)), step=0.5, key="rev_t")
    with c4:
        rev_wellx = st.number_input("WellX ", value=float(quote.get("commission_wellx", 4)), step=0.5, key="rev_w")
    with c5:
        rev_margins = st.number_input("Margins ", value=float(quote.get("commission_margins", 7)), step=0.5, key="rev_m")

    if st.button("📊 Recalculate Premium", type="primary", use_container_width=True):
        burning_cost = float(quote.get("burning_cost", 0) or 0)

        # Recalculate
        projected_claims = burning_cost * 12 * new_census

        # Apply manual adjustment
        net_premium = projected_claims * (1 + manual_adjustment_pct / 100)

        total_comm = (rev_broker + rev_insurer + rev_tpa + rev_wellx + rev_margins) / 100

        if total_comm < 1:
            new_premium = net_premium / (1 - total_comm)
        else:
            new_premium = net_premium
            st.warning("Total commissions >= 100%!")

        st.markdown('<div class="section-lbl">Revised Results</div>', unsafe_allow_html=True)

        col1, col2, col3, col4 = st.columns(4)
        with col1:
            render_metric("Burning Cost Per Capita", burning_cost)
        with col2:
            render_metric("New Census", new_census, currency=False)
        with col3:
            render_metric("Projected Claims", projected_claims)
        with col4:
            st.markdown(f"""
            <div class="premium-hero">
                <div class="ph-label">Revised Premium</div>
                <div class="ph-value">AED {new_premium:,.2f}</div>
                <div class="ph-sub">Per member/year: AED {new_premium / new_census:,.2f}</div>
                <div class="ph-sub">Per member/month: AED {new_premium / new_census / 12:,.2f}</div>
            </div>
            """, unsafe_allow_html=True)

        # Comparison with original
        original_premium = float(quote.get("indicative_premium", 0) or 0)
        if original_premium > 0:
            diff = new_premium - original_premium
            diff_pct = (diff / original_premium) * 100
            badge = "badge-ok" if diff <= 0 else "badge-warn"
            st.markdown(f'<div class="{badge}">Change from original: AED {diff:+,.2f} ({diff_pct:+.1f}%)</div>', unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# PAGE: DASHBOARD
# ---------------------------------------------------------------------------

def page_dashboard():
    """Show all quotes with filtering and status editing."""
    st.markdown("""
    <div style="background:#000; border-radius:14px; overflow:hidden; margin-bottom:28px;">
        <div style="height:3px; background:linear-gradient(90deg,#fb9b35,#f1517b,#b43082,#8431cb,#35c5fc);"></div>
        <div style="padding:22px 28px;">
            <div style="font-family:'Raleway',sans-serif; font-weight:900; font-size:0.6rem;
                        letter-spacing:3px; text-transform:uppercase;
                        background:linear-gradient(90deg,#fb9b35,#f1517b,#b43082,#8431cb,#35c5fc);
                        -webkit-background-clip:text; -webkit-text-fill-color:transparent;
                        margin-bottom:6px;">WellX</div>
            <div style="font-family:'Raleway',sans-serif; font-weight:800; font-size:1.3rem;
                        color:#fff; line-height:1.2; margin-bottom:4px;">
                Dashboard
                <span style="display:block; font-size:0.85rem; font-weight:500;
                             color:rgba(255,255,255,0.45); font-family:'Inter',sans-serif; margin-top:4px;">
                    View, filter, and manage all saved quotes
                </span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    quotes = get_all_quotes()

    if not quotes:
        st.info("No quotes saved yet. Go to 'New Quote' to create one.")
        return

    # Status filter
    all_statuses = ["All", "neutral", "positive", "not good", "will confirm", "confirmed", "lost"]
    status_filter = st.selectbox("Filter by Status", all_statuses)

    if status_filter != "All":
        quotes = [q for q in quotes if q.get("status") == status_filter]

    st.markdown(f"**Showing {len(quotes)} quote(s)**")
    st.markdown("---")

    # Status color badges
    status_colors = {
        "neutral": "🟡",
        "positive": "🟢",
        "not good": "🔴",
        "will confirm": "🟠",
        "confirmed": "✅",
        "lost": "⚫",
    }

    for q in quotes:
        badge = status_colors.get(q["status"], "⚪")
        premium = float(q.get("indicative_premium", 0) or 0)
        census = int(q.get("current_census", 0) or 0)

        with st.expander(
            f"{badge} #{q['id']} — **{q['company_name']}** | "
            f"Premium: AED {premium:,.0f} | Census: {census} | "
            f"{q['created_at'][:10]}"
        ):
            col1, col2, col3 = st.columns([2, 1, 1])
            with col1:
                st.markdown(f"**Broker:** {q.get('broker_name', 'N/A')}")
                st.markdown(f"**Created:** {q['created_at'][:19]}")
                st.markdown(f"**Burning Cost Per Capita:** AED {float(q.get('burning_cost', 0) or 0):,.2f}")

                # Commission breakdown
                comms = f"Broker: {q.get('commission_broker', 0)}% | Insurer: {q.get('commission_insurer', 0)}% | TPA: {q.get('commission_tpa', 0)}% | WellX: {q.get('commission_wellx', 0)}% | Margins: {q.get('commission_margins', 0)}%"
                st.markdown(f"**Commissions:** {comms}")

            with col2:
                new_status = st.selectbox(
                    "Update Status",
                    ["neutral", "positive", "not good", "will confirm", "confirmed", "lost"],
                    index=["neutral", "positive", "not good", "will confirm", "confirmed", "lost"].index(q["status"]),
                    key=f"status_{q['id']}",
                )

            with col3:
                if st.button("Update", key=f"update_{q['id']}", use_container_width=True):
                    update_quote_status(q["id"], new_status)
                    st.success(f"Status updated to '{new_status}'")
                    st.rerun()

            # Show stored summary if available
            if q.get("summary_json"):
                try:
                    stored = json.loads(q["summary_json"])
                    if stored.get("flags"):
                        st.markdown("**Flags:**")
                        for flag in stored["flags"]:
                            st.markdown(f'<div class="badge-warn">⚠ {flag}</div>', unsafe_allow_html=True)
                except (json.JSONDecodeError, TypeError):
                    pass


# ---------------------------------------------------------------------------
# MAIN APP ENTRY POINT
# ---------------------------------------------------------------------------

def main():
    """Main application entry point."""
    setup_page()
    page = render_sidebar()

    if page == "📝 New Quote":
        page_new_quote()
    elif page == "📋 Extracted Information":
        page_extracted_info()
    elif page == "🔄 Revisions":
        page_revisions()
    elif page == "📊 Dashboard":
        page_dashboard()


if __name__ == "__main__":
    main()
